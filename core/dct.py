"""Stage 3 — DCT: Statement × TRX → TRX with rates + Summary."""
from __future__ import annotations

import re
from dataclasses import dataclass
from datetime import datetime, timedelta
from pathlib import Path
from typing import Callable, Optional

import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


LogFn = Callable[[str], None]


@dataclass
class DCTResult:
    total_rows: int
    matched_rows: int
    rates_count: int
    out_path: str


SA_TO_STMT = {
    "SA-EUR_A":    "Statement EUR_A",
    "SA-EUR_MD":   "Statement EUR_MD",
    "SA-EUR_FVuk": "Statement EUR_FVuk",
    "SA-EUR_FVcy": "Statement EUR_FVcy",
    "SA-USD_FVcy": "Statement USD_FVcy",
}

SA_TO_PATH = {
    "SA-EUR_A":    "DCT.AUREXA",
    "SA-EUR_MD":   "DCT.MERTERDATA",
    "SA-EUR_FVuk": "DCT.FUSNVIBEUK",
    "SA-EUR_FVcy": "DCTCY.FUSIOVIBES",
    "SA-USD_FVcy": "DCTCY.FUSIOVIBES",
}

# ─── Styles ───────────────────────────────────────────────────────────────────
HDR_FILL  = PatternFill("solid", start_color="1E293B")
HDR_FONT  = Font(name="Calibri", size=9, bold=True, color="F8FAFC")
ROW_ODD   = PatternFill("solid", start_color="FFFFFF")
ROW_EVEN  = PatternFill("solid", start_color="F8FAFC")
ACC_ODD   = PatternFill("solid", start_color="F0FDF4")
ACC_EVEN  = PatternFill("solid", start_color="DCFCE7")
ACC_HDR   = PatternFill("solid", start_color="166534")
ACC_HFONT = Font(name="Calibri", size=9, bold=True, color="FFFFFF")
SUB_FILL  = PatternFill("solid", start_color="E0F2FE")
DARK_SLATE = PatternFill("solid", start_color="1E293B")

DATA_FONT       = Font(name="Calibri", size=9, color="1E293B")
DATA_FONT_MUTED = Font(name="Calibri", size=9, color="64748B")

HAIR = Side(style="hair", color="E2E8F0")
FULL_BORDER = Border(left=HAIR, right=HAIR, top=HAIR, bottom=HAIR)

RIGHT_COLS  = {"Transaction amount", "Fee", "Rate", "Trn Amount EUR", "Fee EUR"}
CENTER_COLS = {"Currency", "Payment system (VISA/MC)", "Shipment date"}


def _sc(ws, row: int, col: int, val, *, fill=None, font=None,
        align: str = "left", fmt: str | None = None):
    cell = ws.cell(row=row, column=col, value=val)
    if fill:
        cell.fill = fill
    if font:
        cell.font = font
    cell.border = FULL_BORDER
    cell.alignment = Alignment(horizontal=align, vertical="center",
                               wrap_text=False, shrink_to_fit=False)
    if fmt:
        cell.number_format = fmt
    return cell


def _parse_settle_dates(raw: str) -> list[str]:
    raw = str(raw).strip()
    m = re.match(r"(\d{2})\.(\d{2})\.-?(\d{2})\.(\d{2})\.(\d{4})", raw)
    if m:
        d1, mo1, d2, mo2, yr = m.groups()
        start = datetime(int(yr), int(mo1), int(d1))
        end   = datetime(int(yr), int(mo2), int(d2))
        dates, cur = [], start
        while cur <= end:
            dates.append(cur.strftime("%Y-%m-%d"))
            cur += timedelta(days=1)
        return dates
    m2 = re.match(r"(\d{2})\.(\d{2})\.(\d{4})", raw)
    if m2:
        d, mo, yr = m2.groups()
        return [datetime(int(yr), int(mo), int(d)).strftime("%Y-%m-%d")]
    return [raw]


def _extract_rates(stmt_path: str, log: Optional[LogFn] = None) -> dict:
    wb = load_workbook(stmt_path, data_only=True)
    rates: dict = {}

    sa_sheets = [s for s in wb.sheetnames if s.startswith("SA-")]
    for sa_sheet in sa_sheets:
        stmt_sheet = SA_TO_STMT.get(sa_sheet)
        merchant_path = SA_TO_PATH.get(sa_sheet)
        if not stmt_sheet or not merchant_path or stmt_sheet not in wb.sheetnames:
            continue

        ws_stmt = wb[stmt_sheet]
        settle_dates: list[str] = []
        for row in ws_stmt.iter_rows(max_row=15, values_only=True):
            if row[4] and "SETTLEMENT PERIOD" in str(row[4]):
                settle_dates = _parse_settle_dates(str(row[5]).strip())
                break

        if not settle_dates:
            if log:
                log(f"  No date in {stmt_sheet}")
            continue

        ws_sa = wb[sa_sheet]
        sa_rows = list(ws_sa.iter_rows(values_only=True))
        hdrs = [str(c or "").strip() for c in sa_rows[0]]

        def col_idx(name: str) -> int | None:
            try:
                return hdrs.index(name)
            except ValueError:
                return None

        i_amt      = col_idx("Amount")
        i_orig_amt = col_idx("OriginalAmount")
        i_orig_ccy = col_idx("OriginalCCY")

        sums: dict = {}
        for r in sa_rows[1:]:
            if not any(v for v in r):
                continue
            desc = str(r[1] or "")
            if "Settlement with merchants" not in desc:
                continue
            ips = "VISA" if "VISA" in desc else ("MC" if ("MC" in desc or "Mastercard" in desc) else None)
            if not ips:
                continue
            try:
                amount = float(r[i_amt] or 0) if i_amt is not None else 0
            except (TypeError, ValueError):
                continue

            orig_ccy = str(r[i_orig_ccy] or "").strip() if i_orig_ccy is not None else ""
            if not orig_ccy or orig_ccy in ("None", ""):
                orig_ccy = "EUR" if "EUR" in sa_sheet else "USD"
            try:
                orig_amt = float(r[i_orig_amt] or 0) if i_orig_amt is not None else 0
            except (TypeError, ValueError):
                orig_amt = 0
            if orig_amt == 0:
                continue

            key = (ips, orig_ccy)
            sums.setdefault(key, {"amount": 0.0, "orig": 0.0})
            sums[key]["amount"] += amount
            sums[key]["orig"]   += orig_amt

        for (ips, orig_ccy), v in sums.items():
            if v["orig"] != 0:
                rate = v["amount"] / v["orig"]
                for sd in settle_dates:
                    rates[(merchant_path, ips, orig_ccy, sd)] = rate
                if log:
                    log(f"  {merchant_path} | {ips} | {orig_ccy} | {settle_dates[0]}… → {rate:.6f}")

        if not sums and "EUR" in sa_sheet:
            for ips in ["VISA", "MC"]:
                for sd in settle_dates:
                    rates[(merchant_path, ips, "EUR", sd)] = 1.0
            if log:
                log(f"  {merchant_path} | EUR same-ccy → 1.000000")

        # USD→EUR conversion from "X USD -- Y EUR" line
        check_sheet = stmt_sheet
        if "USD" in stmt_sheet:
            eur_equiv = stmt_sheet.replace("USD", "EUR")
            if eur_equiv in wb.sheetnames:
                check_sheet = eur_equiv

        if check_sheet in wb.sheetnames:
            ws_check = wb[check_sheet]
            for row2 in ws_check.iter_rows(max_row=80, values_only=True):
                cell_val = str(row2[2] or "")
                m = re.search(r"([\d,\.]+)\s*USD\s*--\s*([\d,\.]+)\s*EUR", cell_val)
                if m:
                    try:
                        usd_sum = float(m.group(1).replace(",", ""))
                        eur_sum = float(m.group(2).replace(",", ""))
                        if usd_sum != 0:
                            usd_rate = eur_sum / usd_sum
                            for ips in ["VISA", "MC"]:
                                for sd in settle_dates:
                                    rates[(merchant_path, ips, "USD", sd)] = usd_rate
                            if log:
                                log(f"  {merchant_path} | USD→EUR → {usd_rate:.6f}")
                    except (ValueError, TypeError):
                        pass
                    break

    return rates


OUT_COLS = [
    ("Shipment date",                   "Shipment date"),
    ("Merchant path",                   "Merchant path"),
    ("Merchant name",                   "Merchant name"),
    ("ARN",                             "ARN"),
    ("Retention reference nr",          "Retention reference nr"),
    ("Currency",                        "Currency"),
    ("Transaction amount",              "Transaction amount"),
    ("Transaction Turnover Fee Amount", "Fee"),
    ("Payment system (VISA/MC)",        "Payment system (VISA/MC)"),
]
NEW_COLS    = ["Rate", "Trn Amount EUR", "Fee EUR"]
NEW_COL_SET = set(NEW_COLS)

# Payment ID is enriched from Stage 1 output — styled separately
PID_FILL_ODD  = PatternFill("solid", start_color="EEF2FF")   # soft indigo
PID_FILL_EVEN = PatternFill("solid", start_color="E0E7FF")
PID_HDR_FILL  = PatternFill("solid", start_color="4338CA")   # indigo header
PID_HDR_FONT  = Font(name="Calibri", size=9, bold=True, color="FFFFFF")
PID_FONT      = Font(name="Calibri", size=9, color="312E81")

COL_WIDTHS = {
    "Shipment date": 13, "Merchant path": 20, "Merchant name": 22,
    "ARN": 26, "Retention reference nr": 20, "Currency": 9,
    "Transaction amount": 16, "Fee": 12, "Payment system (VISA/MC)": 18,
    "Rate": 14, "Trn Amount EUR": 16, "Fee EUR": 13,
    "Payment ID": 32,
}


def _norm_arn(val) -> str:
    """Normalize ARN: strip whitespace, remove .0 float suffix."""
    return str(val or "").split(".")[0].strip()


def _load_arn_payment_id_map(recon_path: str, log: Optional[LogFn] = None) -> dict[str, str]:
    """Read DCT Stage 1 output → {ARN: Payment ID}.

    Tries 'New Transactions' sheet first, then all sheets until ARN + Payment ID
    columns are found.
    """
    try:
        wb = load_workbook(recon_path, read_only=True, data_only=True)

        # Try preferred sheet first, then all others
        preferred = ["New Transactions", "SPNT"]
        candidates = preferred + [s for s in wb.sheetnames if s not in preferred]

        for sheet in candidates:
            if sheet not in wb.sheetnames:
                continue
            ws = wb[sheet]
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                continue
            headers = [str(c or "").strip() for c in rows[0]]

            # Case-insensitive column search
            headers_lower = [h.lower() for h in headers]

            def col_idx(name: str) -> int | None:
                try:
                    return headers_lower.index(name.lower())
                except ValueError:
                    return None

            arn_idx = col_idx("ARN")
            pid_idx = col_idx("Payment ID")

            if arn_idx is None or pid_idx is None:
                if log:
                    log(f"  Sheet '{sheet}': no ARN/Payment ID columns — skipping")
                continue

            mapping: dict[str, str] = {}
            for row in rows[1:]:
                arn = _norm_arn(row[arn_idx])
                pid = str(row[pid_idx] or "").strip()
                if arn and pid and arn not in ("None", "nan") and pid not in ("None", "nan"):
                    mapping[arn] = pid

            if log:
                log(f"  Sheet '{sheet}': {len(mapping)} ARN→Payment ID entries loaded")
            # Show a few samples in log for debugging
            if log and mapping:
                sample = list(mapping.items())[:3]
                log(f"  Sample: {sample}")
            return mapping

        if log:
            log(f"  No suitable sheet found in Stage 1 file (sheets: {wb.sheetnames})")
        return {}
    except Exception as exc:
        if log:
            log(f"  Could not read Stage 1 file: {exc}")
        return {}


def _build_workbook(rates: dict, trx_path: str, out_path: str,
                    arn_pid_map: dict[str, str] | None = None,
                    log: Optional[LogFn] = None) -> tuple[int, int]:
    wb_trx = load_workbook(trx_path)
    ws = wb_trx.active
    rows = list(ws.iter_rows(values_only=True))
    headers = list(rows[0])

    def ci(name: str) -> int | None:
        try:
            return headers.index(name)
        except ValueError:
            return None

    path_idx = ci("Merchant path")
    ips_idx  = ci("Payment system (VISA/MC)")
    ccy_idx  = ci("Currency")
    ship_idx = ci("Shipment date")
    amt_idx  = ci("Transaction amount")
    fee_idx  = ci("Transaction Turnover Fee Amount")

    wb_out = Workbook()
    ws_out = wb_out.active
    ws_out.title = "TRX with Rates"

    src_indices, out_names = [], []
    for src_name, out_name in OUT_COLS:
        idx = ci(src_name)
        src_indices.append(idx)
        out_names.append(out_name)

    # Column appears whenever a recon file was provided (even if 0 matches)
    has_pid = arn_pid_map is not None
    # Payment ID column is inserted right after ARN (index 3, 0-based → position 4, 1-based)
    ARN_OUTPUT_POS = next(
        (i for i, (_, name) in enumerate(zip(src_indices, out_names)) if name == "ARN"), None
    )
    pid_insert_after = (ARN_OUTPUT_POS + 1) if ARN_OUTPUT_POS is not None else len(out_names)

    # Build final header list with optional Payment ID injected after ARN
    def _build_header_list() -> list[str]:
        names = list(out_names)
        if has_pid:
            names.insert(pid_insert_after, "Payment ID")
        return names + NEW_COLS

    all_headers = _build_header_list()
    ws_out.row_dimensions[1].height = 22

    for ci_, h in enumerate(all_headers, 1):
        is_pid = has_pid and h == "Payment ID"
        is_acc = h in NEW_COL_SET
        al = "right" if h in RIGHT_COLS else ("center" if h in CENTER_COLS else "left")
        if is_pid:
            fill, font = PID_HDR_FILL, PID_HDR_FONT
        elif is_acc:
            fill, font = ACC_HDR, ACC_HFONT
        else:
            fill, font = HDR_FILL, HDR_FONT
        _sc(ws_out, 1, ci_, h, fill=fill, font=font, align=al)
        ws_out.column_dimensions[get_column_letter(ci_)].width = COL_WIDTHS.get(h, 14)

    ws_out.freeze_panes = "A2"

    data_rows = [
        r for r in rows[1:]
        if any(v for v in r) and
        not all(str(v or "").strip() in ("", "'_", "'--") for v in r)
    ]

    matched = 0
    summary: dict = {}

    for ri, row in enumerate(data_rows, 2):
        even = ri % 2 == 0
        base = ROW_EVEN if even else ROW_ODD
        acc  = ACC_EVEN if even else ACC_ODD

        path     = str(row[path_idx] or "") if path_idx is not None else ""
        ips      = str(row[ips_idx]  or "") if ips_idx  is not None else ""
        ccy      = str(row[ccy_idx]  or "") if ccy_idx  is not None else ""
        ship     = row[ship_idx] if ship_idx is not None else None
        ship_str = str(ship)[:10] if ship else ""

        rate = rates.get((path, ips, ccy, ship_str))
        if rate is not None:
            matched += 1

        try:
            trn_amt = float(row[amt_idx] or 0) if amt_idx is not None else None
        except (TypeError, ValueError):
            trn_amt = None
        try:
            fee_amt = float(row[fee_idx] or 0) if fee_idx is not None else None
        except (TypeError, ValueError):
            fee_amt = None

        trn_eur = round(trn_amt * rate, 6) if rate and trn_amt is not None else None
        fee_eur = round(fee_amt * rate, 6) if rate and fee_amt is not None else None

        # Resolve Payment ID from ARN
        arn_raw = _norm_arn(row[ci("ARN")]) if ci("ARN") is not None else ""
        payment_id = (arn_pid_map or {}).get(arn_raw, "")

        # Write base columns, injecting Payment ID after ARN
        col_cursor = 1
        for i, (idx, hname) in enumerate(zip(src_indices, out_names)):
            val = row[idx] if idx is not None else None
            fmt = None
            al  = "right" if hname in RIGHT_COLS else ("center" if hname in CENTER_COLS else "left")

            if hname == "Shipment date" and val is not None:
                try:
                    val = val.strftime("%d.%m.%Y") if hasattr(val, "strftime") else str(val)[:10]
                except (AttributeError, ValueError):
                    val = str(val)[:10]
            if hname in ("Transaction amount", "Fee"):
                fmt = "#,##0.00"

            font = DATA_FONT_MUTED if hname == "Merchant path" else DATA_FONT
            _sc(ws_out, ri, col_cursor, val, fill=base, font=font, align=al, fmt=fmt)
            col_cursor += 1

            # Inject Payment ID right after ARN
            if has_pid and i == pid_insert_after - 1:
                pid_fill = PID_FILL_EVEN if even else PID_FILL_ODD
                _sc(ws_out, ri, col_cursor, payment_id or None,
                    fill=pid_fill, font=PID_FONT, align="left")
                col_cursor += 1

        extra = col_cursor - 1
        _sc(ws_out, ri, extra+1, rate,    fill=acc, font=DATA_FONT, align="right", fmt="0.000000")
        _sc(ws_out, ri, extra+2, trn_eur, fill=acc, align="right", fmt="#,##0.00",
            font=Font(name="Calibri", size=9, bold=True, color="166534"))
        _sc(ws_out, ri, extra+3, fee_eur, fill=acc, align="right", fmt="#,##0.00",
            font=Font(name="Calibri", size=9, bold=True, color="991B1B"))

        # Accumulate summary
        p_key = path or "(blank)"
        c_key = ccy  or "(blank)"
        key = (p_key, c_key)
        summary.setdefault(key, {"trn": 0.0, "fee": 0.0, "count": 0})
        summary[key]["count"] += 1
        if rate:
            summary[key]["trn"] += (trn_amt or 0) * rate
            summary[key]["fee"] += (fee_amt or 0) * rate

    # Sheet 2: Summary
    ws_sum = wb_out.create_sheet("Summary")
    ws_sum.row_dimensions[1].height = 22
    SUM_HEADERS = ["Merchant path", "Currency", "Trn Amount EUR", "Fee EUR", "Count"]
    SUM_WIDTHS  = [26, 10, 18, 18, 10]
    SUM_ALIGNS  = ["left", "center", "right", "right", "center"]
    for ci_, (h, w, al) in enumerate(zip(SUM_HEADERS, SUM_WIDTHS, SUM_ALIGNS), 1):
        _sc(ws_sum, 1, ci_, h, fill=HDR_FILL, font=HDR_FONT, align=al)
        ws_sum.column_dimensions[get_column_letter(ci_)].width = w

    sorted_keys = sorted(summary.keys(), key=lambda x: (x[0].lower(), x[1].lower()))
    ri = 2
    prev_path = None
    path_trn = path_fee = path_cnt = 0.0

    def write_subtotal(row_i: int, path_name: str, trn: float, fee: float, cnt: float) -> None:
        for ci_, val in enumerate(
                [path_name + " — Subtotal", "", round(trn, 6) or None, round(fee, 6) or None, int(cnt)], 1):
            fmt = "#,##0.000000" if ci_ in (3, 4) else None
            al  = "right" if ci_ in (3, 4) else ("center" if ci_ == 5 else "left")
            _sc(ws_sum, row_i, ci_, val, fill=SUB_FILL,
                font=Font(name="Calibri", size=9, bold=True, color="0C4A6E"),
                align=al, fmt=fmt)

    for path, ccy in sorted_keys:
        vals = summary[(path, ccy)]
        if prev_path is not None and path != prev_path:
            write_subtotal(ri, prev_path, path_trn, path_fee, path_cnt)
            ri += 1
            path_trn = path_fee = path_cnt = 0.0

        prev_path = path
        path_trn  += vals["trn"]
        path_fee  += vals["fee"]
        path_cnt  += vals["count"]

        even = ri % 2 == 0
        _sc(ws_sum, ri, 1, path, fill=ROW_EVEN if even else ROW_ODD, font=DATA_FONT_MUTED)
        _sc(ws_sum, ri, 2, ccy,  fill=ROW_EVEN if even else ROW_ODD, font=DATA_FONT, align="center")
        _sc(ws_sum, ri, 3, round(vals["trn"], 2) or None,
            fill=ACC_EVEN if even else ACC_ODD, align="right", fmt="#,##0.00",
            font=Font(name="Calibri", size=9, bold=True, color="166534"))
        _sc(ws_sum, ri, 4, round(vals["fee"], 2) or None,
            fill=ACC_EVEN if even else ACC_ODD, align="right", fmt="#,##0.00",
            font=Font(name="Calibri", size=9, bold=True, color="991B1B"))
        _sc(ws_sum, ri, 5, vals["count"], fill=ROW_EVEN if even else ROW_ODD, font=DATA_FONT, align="center")
        ri += 1

    if prev_path:
        write_subtotal(ri, prev_path, path_trn, path_fee, path_cnt)
        ri += 1

    total_trn = sum(v["trn"] for v in summary.values())
    total_fee = sum(v["fee"] for v in summary.values())
    total_cnt = sum(v["count"] for v in summary.values())
    _sc(ws_sum, ri, 1, "Grand Total", fill=DARK_SLATE, font=Font(name="Calibri", size=9, bold=True, color="F8FAFC"))
    _sc(ws_sum, ri, 2, "", fill=DARK_SLATE, font=HDR_FONT)
    _sc(ws_sum, ri, 3, round(total_trn, 6), fill=DARK_SLATE, align="right", fmt="#,##0.00",
        font=Font(name="Calibri", size=9, bold=True, color="86EFAC"))
    _sc(ws_sum, ri, 4, round(total_fee, 6), fill=DARK_SLATE, align="right", fmt="#,##0.00",
        font=Font(name="Calibri", size=9, bold=True, color="FCA5A5"))
    _sc(ws_sum, ri, 5, int(total_cnt), fill=DARK_SLATE, align="center",
        font=Font(name="Calibri", size=9, bold=True, color="F8FAFC"))

    ws_sum.freeze_panes = "A2"
    wb_out.save(out_path)
    return len(data_rows), matched


def run_dct(*, statement_path: str, trx_path: str, out_dir: str,
            recon_path: Optional[str] = None,
            log: Optional[LogFn] = None) -> DCTResult:
    def lg(msg: str) -> None:
        if log:
            log(msg)

    lg("Extracting rates from Statement…")
    rates = _extract_rates(statement_path, lg)
    lg(f"  Rate keys found: {len(rates)}")

    arn_pid_map: dict[str, str] | None = None
    if recon_path:
        lg("Loading ARN → Payment ID from Stage 1 output…")
        arn_pid_map = _load_arn_payment_id_map(recon_path, lg)

    lg("Reading TRX file…")
    out_name = f"TRX_with_rates_{datetime.now().strftime('%d-%m-%Y')}.xlsx"
    out_path = str(Path(out_dir) / out_name)

    lg("Building workbook…")
    total, matched = _build_workbook(rates, trx_path, out_path, arn_pid_map, lg)
    lg(f"  Matched: {matched} / {total}")

    return DCTResult(
        total_rows=total,
        matched_rows=matched,
        rates_count=len(rates),
        out_path=out_path,
    )
