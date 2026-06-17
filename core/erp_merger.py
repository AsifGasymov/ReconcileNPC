"""CDQ Merger: Transactions × Statement (FX) × Orders → ERP Upload."""
from __future__ import annotations

import re
from dataclasses import dataclass
from datetime import datetime
from decimal import Decimal, ROUND_HALF_UP
from pathlib import Path
from typing import Callable, Optional

import openpyxl
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

from .excel_styles import (
    ALT, DARK_BLUE, GOLD, GOLD_ALT, GREEN_FILL, MID_BLUE, ORANGE_ALT, ORANGE_FILL, WHITE,
    data_font, header_font, style_cell,
)


LogFn = Callable[[str], None]

_CENT = Decimal("0.01")
_RATE5 = Decimal("0.00001")


def _br(value, *, places: int = 2) -> float:
    """Banker's rounding (ROUND_HALF_UP) to `places` decimal digits."""
    exp = _CENT if places == 2 else Decimal("0." + "0" * places)
    return float(Decimal(str(value)).quantize(exp, rounding=ROUND_HALF_UP))


@dataclass
class ERPResult:
    rows: int
    rates_count: int
    total_gbp: float
    total_eur: float
    out_path: str
    out_paths: list[str]


# EU Fee / Non-EU Fee are in original currency; Approve Fee is always in EUR
FEE_ORIGINAL_COLS = ["EU Fee", "Non-EU Fee"]
FEE_EUR_COL = "Approve Fee"

OUTPUT_COLS: list[tuple[str, int]] = [
    ("Company", 18), ("IPS", 7), ("Region", 10),
    ("RRN", 16), ("Payment ID", 32),
    ("Statement Date", 16),
    ("Transaction Date/Time", 20), ("Processing Date/Time", 20),
    ("Original currency", 12), ("Original amount", 14),
    ("Amount in GBP", 16), ("Currency", 8), ("Amount in EUR", 16),
    ("Fee", 14), ("Approve Fee", 14),
]

GOLD_COLS = {"Amount in GBP", "Amount in EUR"}
TEXT_COLS = {"RRN"}
NUMBER_FORMATS = {
    "Original amount": "#,##0.00",
    "Amount in GBP": "#,##0.00",
    "Amount in EUR": "#,##0.00",
    "Fee": "#,##0.0000",
    "Approve Fee": "#,##0.00",
}


def _sanitize_filename(name: str) -> str:
    name = re.sub(r'[\\/:*?"<>|]', '', name)
    return name.strip()[:40]


def _is_priority_row(pid) -> bool:
    s = str(pid).strip() if pd.notna(pid) else ""
    return s == "" or s.lower().startswith("charge for")


def _extract_rates(stmt_path: str) -> tuple[dict[tuple[str, str], float], str]:
    wb = openpyxl.load_workbook(stmt_path, data_only=True)
    rates: dict[tuple[str, str], float] = {}
    stmt_date = ""

    for sheet_name in wb.sheetnames:
        ccy = sheet_name.upper()
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))

        if not stmt_date:
            for row in rows[:6]:
                for cell in row:
                    s = str(cell or "")
                    if re.match(r"\d{2}\.\d{2}\.\d{4}", s):
                        stmt_date = s
                        break
                if stmt_date:
                    break

        # GBP and EUR are settlement currencies — rate to GBP is always 1
        if ccy in ("GBP", "EUR"):
            rates[(ccy, "MC")] = 1.0
            rates[(ccy, "VISA")] = 1.0
            continue

        mc_rate = visa_rate = None
        for row in rows:
            a0 = str(row[0] or "")
            a2 = str(row[2] or "").replace(",", ".")
            if "Mastercard Payout" in a0:
                try:
                    mc_rate = float(a2)
                except ValueError:
                    pass
            if "Visa Payout" in a0:
                try:
                    visa_rate = float(a2)
                except ValueError:
                    pass

        if mc_rate is not None:
            rates[(ccy, "MC")] = mc_rate
        if visa_rate is not None:
            rates[(ccy, "VISA")] = visa_rate

    return rates, stmt_date


def _read_transactions(trn_path: str) -> pd.DataFrame:
    p = trn_path.lower()
    if p.endswith(".csv"):
        try:
            df = pd.read_csv(trn_path, encoding="utf-8")
        except UnicodeDecodeError:
            df = pd.read_csv(trn_path, encoding="latin-1")
    else:
        df = pd.read_excel(trn_path)
    return df.dropna(how="all")


def _read_orders(orders_path: str) -> pd.DataFrame:
    df = pd.read_excel(orders_path, sheet_name="Orders",
                       dtype={"ARN": str, "RRN": str})
    df = df.dropna(how="all")
    df["ARN"] = df["ARN"].apply(
        lambda x: str(x).split(".")[0].strip() if pd.notna(x) else "")
    df["RRN"] = df["RRN"].apply(
        lambda x: str(x).split(".")[0].strip() if pd.notna(x) else "")
    df["Payment ID"] = df["Product name or description"].apply(
        lambda x: str(x)[:28].strip() if pd.notna(x) else "")
    return df[["ARN", "RRN", "Payment ID"]]


def _build_workbook(trn_df: pd.DataFrame, rates: dict, stmt_date: str,
                    out_path: str, orders_df: Optional[pd.DataFrame],
                    eur_rate: float) -> tuple[int, int, float, float]:
    trn_df = trn_df.copy()

    # Normalise column names to match OUTPUT_COLS
    trn_df.rename(columns={
        "Currency": "Original currency",
        "Trn Amount": "Original amount",
    }, inplace=True)

    if orders_df is not None and not orders_df.empty:
        trn_df["_RRN"] = trn_df["RRN"].apply(
            lambda x: str(x).split(".")[0].strip() if pd.notna(x) else "")
        rrn_map = orders_df[orders_df["RRN"] != ""].set_index("RRN")["Payment ID"].to_dict()
        trn_df["Payment ID"] = trn_df["_RRN"].map(rrn_map).fillna("")
        trn_df.drop(columns=["_RRN"], inplace=True)
    else:
        trn_df["Payment ID"] = ""

    trn_df["Statement Date"] = stmt_date

    def _fx_to_gbp(row):
        ccy = str(row.get("Original currency", "")).upper().strip()
        if ccy in ("GBP", "EUR"):
            return 1.0
        ips = str(row.get("IPS", "")).upper().strip()
        return rates.get((ccy, ips))

    trn_df["_fx_rate"] = trn_df.apply(_fx_to_gbp, axis=1)
    trn_df["Original amount"] = pd.to_numeric(trn_df["Original amount"], errors="coerce").fillna(0)

    # Normalise fee source columns to numeric
    for c in FEE_ORIGINAL_COLS + [FEE_EUR_COL]:
        if c in trn_df.columns:
            trn_df[c] = pd.to_numeric(trn_df[c], errors="coerce").fillna(0)

    def _compute_fee(row):
        orig_fees = sum(float(row.get(c, 0) or 0) for c in FEE_ORIGINAL_COLS if c in trn_df.columns)
        approve_eur = float(row.get(FEE_EUR_COL, 0) or 0) if FEE_EUR_COL in trn_df.columns else 0.0

        ccy = str(row.get("Original currency", "")).upper().strip()
        fx = row.get("_fx_rate")

        if ccy == "EUR":
            # EU Fee + Non-EU Fee are already in EUR — no conversion needed
            fee_eur = _br(float(orig_fees))
        elif fx is not None and pd.notna(fx) and float(fx) != 0:
            # original ccy → GBP → EUR, banker's rounding at each step
            fee_eur = _br(_br(float(orig_fees) * float(fx)) * eur_rate)
        else:
            fee_eur = _br(float(orig_fees))  # no rate available, add as-is

        return _br(fee_eur + approve_eur, places=4)

    trn_df["Fee"] = trn_df.apply(_compute_fee, axis=1)

    # Expose Approve Fee as its own column (already in EUR, normalised above)
    if FEE_EUR_COL in trn_df.columns:
        trn_df["Approve Fee"] = trn_df[FEE_EUR_COL]
    else:
        trn_df["Approve Fee"] = 0.0

    trn_df["Amount in GBP"] = trn_df.apply(
        lambda r: _br(float(r["Original amount"]) * float(r["_fx_rate"]))
        if pd.notna(r["_fx_rate"]) else None, axis=1)

    def _eur_amount(row):
        ccy = str(row.get("Original currency", "")).upper().strip()
        orig = row.get("Original amount", 0)
        if ccy == "EUR":
            return _br(float(orig))
        gbp = row.get("Amount in GBP")
        if gbp is None or (not isinstance(gbp, str) and pd.isna(gbp)):
            return None
        return _br(float(gbp) * eur_rate)

    trn_df["Amount in EUR"] = trn_df.apply(_eur_amount, axis=1)
    trn_df["Currency"] = "EUR"

    trn_df["_priority"] = trn_df["Payment ID"].apply(_is_priority_row)
    trn_df = trn_df.sort_values("_priority", ascending=False).reset_index(drop=True)

    wb = Workbook()

    # ── Sheet 1: ERP Upload ───────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "ERP Upload"
    ws1.row_dimensions[1].height = 18

    for ci, (name, width) in enumerate(OUTPUT_COLS, 1):
        style_cell(ws1, 1, ci, name, fill=MID_BLUE,
                   font=header_font(size=9), align="center")
        ws1.column_dimensions[get_column_letter(ci)].width = width

    src_cols = [c[0] for c in OUTPUT_COLS]
    for ri, (_, row) in enumerate(trn_df.iterrows(), 2):
        even = ri % 2 == 0
        is_priority = bool(row.get("_priority", False))
        for ci, col in enumerate(src_cols, 1):
            val = row.get(col)
            if not isinstance(val, str) and pd.isna(val):
                val = None
            if col in TEXT_COLS and val is not None:
                val = str(val).split(".")[0]
            is_gold = col in GOLD_COLS
            if is_priority:
                fill = ORANGE_FILL if even else ORANGE_ALT
            elif is_gold:
                fill = GOLD if even else GOLD_ALT
            else:
                fill = ALT if even else WHITE
            cell = style_cell(ws1, ri, ci, val, fill=fill,
                              font=data_font(),
                              number_format=NUMBER_FORMATS.get(col))
            if col in TEXT_COLS:
                cell.number_format = "@"

    ws1.freeze_panes = "A2"
    trn_df.drop(columns=["_priority", "_fx_rate"], inplace=True, errors="ignore")

    # ── Sheet 2: FX Rates ─────────────────────────────────────────────────────
    ws2 = wb.create_sheet("FX Rates")
    ws2.merge_cells("A1:E1")
    fx_title = ws2["A1"]
    fx_title.value = f"FX Rates Applied   |   Source: Statement {stmt_date}"
    fx_title.font = Font(name="Arial", size=12, bold=True, color="FFFFFF")
    fx_title.fill = DARK_BLUE
    fx_title.alignment = Alignment(horizontal="center", vertical="center")
    ws2.row_dimensions[1].height = 26

    for ci, (h, w) in enumerate(
            [("Currency", 14), ("IPS", 10), ("Rate to GBP", 16), ("Note", 40)], 1):
        style_cell(ws2, 2, ci, h, fill=MID_BLUE,
                   font=header_font(size=9), align="center")
        ws2.column_dimensions[get_column_letter(ci)].width = w

    note_map = {
        ("GBP", "MC"): "Base currency", ("GBP", "VISA"): "Base currency",
        ("EUR", "MC"): "Settlement currency", ("EUR", "VISA"): "Settlement currency",
    }
    sorted_rates = sorted(rates.items())
    for ri, ((ccy, ips), rate) in enumerate(sorted_rates, 3):
        fill = ALT if ri % 2 == 0 else WHITE
        style_cell(ws2, ri, 1, ccy, fill=fill, font=data_font(True))
        style_cell(ws2, ri, 2, ips, fill=fill, font=data_font())
        style_cell(ws2, ri, 3, rate, fill=GOLD, font=data_font(),
                   number_format="0.00000", align="center")
        style_cell(ws2, ri, 4,
                   note_map.get((ccy, ips), "From statement Payout row"),
                   fill=fill, font=data_font())

    # EUR rate row (user-defined GBP→EUR)
    eur_ri = 3 + len(sorted_rates)
    fill = ALT if eur_ri % 2 == 0 else WHITE
    style_cell(ws2, eur_ri, 1, "EUR", fill=fill, font=data_font(True))
    style_cell(ws2, eur_ri, 2, "—", fill=fill, font=data_font())
    style_cell(ws2, eur_ri, 3, eur_rate, fill=GREEN_FILL, font=data_font(),
               number_format="0.00000", align="center")
    style_cell(ws2, eur_ri, 4, "User-defined GBP→EUR rate (Amount in EUR column)",
               fill=fill, font=data_font())

    # ── Sheet 3: Summary ──────────────────────────────────────────────────────
    ws3 = wb.create_sheet("Summary")

    # Build grouped aggregation — add fee columns only if present in the file
    fee_present = [c for c in FEE_ORIGINAL_COLS if c in trn_df.columns]
    if fee_present:
        trn_df["_eu_noneu_fee"] = trn_df[fee_present].sum(axis=1)
    else:
        trn_df["_eu_noneu_fee"] = 0.0

    summary = trn_df.groupby(["Original currency", "IPS"]).agg(
        Count=("Original amount", "count"),
        Total_Original=("Original amount", "sum"),
        Total_Fee=("_eu_noneu_fee", "sum"),
        Total_GBP=("Amount in GBP", "sum"),
        Total_EUR=("Amount in EUR", "sum"),
    ).reset_index()

    summary["Net_Original"] = summary["Total_Original"] - summary["Total_Fee"]
    trn_df.drop(columns=["_eu_noneu_fee"], inplace=True, errors="ignore")

    SCOLS = [
        ("Currency", 12), ("IPS", 10), ("Trn Count", 12),
        ("Total Original", 18), ("EU+Non-EU Fee", 18), ("Net Original", 18),
        ("Total GBP", 16), ("Total EUR", 16),
    ]

    ws3.merge_cells(f"A1:{get_column_letter(len(SCOLS))}1")
    sum_title = ws3["A1"]
    sum_title.value = f"Summary by Currency & IPS   |   {stmt_date}"
    sum_title.font = Font(name="Arial", size=12, bold=True, color="FFFFFF")
    sum_title.fill = DARK_BLUE
    sum_title.alignment = Alignment(horizontal="center", vertical="center")
    ws3.row_dimensions[1].height = 26

    for ci, (h, w) in enumerate(SCOLS, 1):
        style_cell(ws3, 2, ci, h, fill=MID_BLUE,
                   font=header_font(size=9), align="center")
        ws3.column_dimensions[get_column_letter(ci)].width = w

    for ri, (_, row) in enumerate(summary.iterrows(), 3):
        fill = ALT if ri % 2 == 0 else WHITE
        style_cell(ws3, ri, 1, row["Original currency"], fill=fill, font=data_font(True))
        style_cell(ws3, ri, 2, row["IPS"], fill=fill, font=data_font())
        style_cell(ws3, ri, 3, int(row["Count"]), fill=fill,
                   font=data_font(), align="center")
        style_cell(ws3, ri, 4, _br(row["Total_Original"]),
                   fill=fill, font=data_font(), number_format="#,##0.00")
        style_cell(ws3, ri, 5, _br(float(row["Total_Fee"] or 0)),
                   fill=ORANGE_FILL, font=data_font(), number_format="#,##0.0000")
        style_cell(ws3, ri, 6, _br(float(row["Net_Original"] or 0)),
                   fill=fill, font=data_font(True), number_format="#,##0.00")
        style_cell(ws3, ri, 7, _br(float(row["Total_GBP"] or 0)),
                   fill=GOLD, font=data_font(True), number_format="#,##0.00")
        style_cell(ws3, ri, 8, _br(float(row["Total_EUR"] or 0)),
                   fill=GREEN_FILL, font=data_font(True), number_format="#,##0.00")

    last = 3 + len(summary)
    ws3.merge_cells(f"A{last}:C{last}")
    style_cell(ws3, last, 1, "TOTAL", fill=DARK_BLUE,
               font=header_font(), align="right")
    style_cell(ws3, last, 4, _br(summary["Total_Original"].sum()),
               fill=DARK_BLUE, font=header_font(), number_format="#,##0.00")
    style_cell(ws3, last, 5, _br(float(summary["Total_Fee"].sum())),
               fill=DARK_BLUE, font=header_font(), number_format="#,##0.0000")
    style_cell(ws3, last, 6, _br(float(summary["Net_Original"].sum())),
               fill=DARK_BLUE, font=header_font(), number_format="#,##0.00")
    style_cell(ws3, last, 7, _br(float(summary["Total_GBP"].sum())),
               fill=DARK_BLUE,
               font=Font(name="Arial", size=10, bold=True, color="FFD700"),
               number_format="#,##0.00")
    style_cell(ws3, last, 8, _br(float(summary["Total_EUR"].sum())),
               fill=DARK_BLUE,
               font=Font(name="Arial", size=10, bold=True, color="10B981"),
               number_format="#,##0.00")

    wb.save(out_path)
    return (
        len(trn_df),
        len(rates),
        float(summary["Total_GBP"].sum()),
        float(summary["Total_EUR"].sum()),
    )


def run_erp_merger(*, transactions_paths: list[str], statement_paths: list[str],
                   out_dir: str, orders_path: Optional[str] = None,
                   eur_rate: float = 1.0,
                   log: Optional[LogFn] = None) -> ERPResult:
    """Merge provider reports into a single ERP-ready workbook per statement."""
    def lg(msg: str) -> None:
        if log:
            log(msg)

    lg(f"EUR rate (GBP→EUR): {eur_rate}")

    orders_df = None
    if orders_path:
        lg("Reading order statement…")
        orders_df = _read_orders(orders_path)

    pairs = list(zip(transactions_paths, statement_paths))
    total_rows = 0
    total_rates = 0
    total_gbp = 0.0
    total_eur = 0.0
    out_paths: list[str] = []

    for i, (trn_path, stmt_path) in enumerate(pairs, 1):
        prefix = f"[{i}/{len(pairs)}] " if len(pairs) > 1 else ""

        lg(f"{prefix}Reading statement…")
        rates, stmt_date = _extract_rates(stmt_path)
        lg(f"  rates: {len(rates)} · date: {stmt_date or 'n/a'}")

        lg(f"{prefix}Reading transactions…")
        trn_df = _read_transactions(trn_path)
        lg(f"  rows: {len(trn_df)}")

        companies = trn_df["Company"].dropna() if "Company" in trn_df.columns else pd.Series(dtype=str)
        companies = companies[companies.astype(str).str.strip().ne("")]
        company = _sanitize_filename(str(companies.mode()[0])) if not companies.empty else "Unknown"

        date_tag = stmt_date.replace(".", "-") if stmt_date else datetime.now().strftime("%d-%m-%Y")
        out_name = f"CDQ_{company}_{date_tag}.xlsx"
        out_path = str(Path(out_dir) / out_name)

        lg(f"{prefix}Writing {out_name}…")
        rows, n_rates, gbp, eur = _build_workbook(
            trn_df, rates, stmt_date, out_path, orders_df, eur_rate)

        total_rows += rows
        total_rates = max(total_rates, n_rates)
        total_gbp += gbp
        total_eur += eur
        out_paths.append(out_path)

    return ERPResult(
        rows=total_rows,
        rates_count=total_rates,
        total_gbp=total_gbp,
        total_eur=total_eur,
        out_path=out_dir,
        out_paths=out_paths,
    )
