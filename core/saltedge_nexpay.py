"""SaltEdge × Nexpay reconciliation report."""
from __future__ import annotations

import os
from dataclasses import dataclass
from datetime import datetime
from typing import Callable, Optional

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from .excel_styles import (
    ALT, BORDER, DARK_BLUE, GREEN_FILL, ORANGE_FILL, WHITE,
    data_font, header_font, style_cell,
)

LogFn = Callable[[str], None]

SALTEDGE_MERCHANT_NEXPAY = "NexPay"

STATUS_MATCHED   = "Matched"
STATUS_SE_ONLY   = "SaltEdge only"
STATUS_NX_ONLY   = "Nexpay only"

NUMBER_FMT = "#,##0.00"

# Sheet 1 & 3 — matched rows (both sides present)
MATCHED_COLS: list[tuple[str, int]] = [
    ("Status",              16),
    ("Payment ID",          32),
    ("Customer Name",       24),
    ("SE Amount (EUR)",     16),
    ("SE Fee",              12),
    ("SE Date",             14),
    ("NX Creation Date",    18),
    ("NX Processing Date",  18),
    ("NX Beneficiary",      24),
    ("NX Account",          24),
    ("NX Amount (EUR)",     16),
    ("NX Payment number",   18),
    ("Difference",          14),
]

# Sheet 2 — SE processed rows missing from Nexpay
SE_PROC_COLS: list[tuple[str, int]] = [
    ("Payment ID",        32),
    ("Customer Name",     24),
    ("Payment status",    16),
    ("SE Amount (EUR)",   16),
    ("SE Fee",            12),
    ("SE Date",           16),
]

# Sheet 3 — NX rows missing from SE
NX_PROC_COLS: list[tuple[str, int]] = [
    ("NX Creation Date",    18),
    ("NX Processing Date",  18),
    ("NX Beneficiary",      24),
    ("NX Account",          24),
    ("NX Amount (EUR)",     16),
    ("NX Payment number",   18),
    ("Details",             40),
]

# Sheet 5 — all unmatched (SE + NX combined)
UNMATCHED_COLS: list[tuple[str, int]] = [
    ("Source",                  14),
    ("Status",                  16),
    ("Payment ID / NX Ref",     32),
    ("Customer / Beneficiary",  28),
    ("Amount (EUR)",            16),
    ("Date",                    16),
    ("Details / Account",       36),
]

AMOUNT_COLS = {"SE Amount (EUR)", "NX Amount (EUR)", "SE Fee", "Difference", "Amount (EUR)"}


@dataclass
class NexpayResult:
    matched: int                  # sheet 1 — all matched
    se_processed_nx_missing: int  # sheet 2 — SE processed, no NX match
    nx_processed_se_missing: int  # sheet 3 — NX rows, no SE match
    matched_exc: int              # sheet 4 — matched but not processed
    unmatched: int                # sheet 5 — SE_ONLY + NX_ONLY
    total_se_amount: float
    total_nx_amount: float
    out_path: str


def run_saltedge_nexpay(
    *,
    system_paths: list[str],
    nexpay_path: str,
    out_dir: str,
    log: Optional[LogFn] = None,
) -> NexpayResult:
    def _log(msg: str) -> None:
        if log:
            log(msg)

    _log(f"Reading {len(system_paths)} system export(s)…")
    parts = []
    for p in system_paths:
        df = pd.read_csv(p, dtype=str)
        df.columns = [c.strip() for c in df.columns]
        parts.append(df)
    se = pd.concat(parts, ignore_index=True) if len(parts) > 1 else parts[0]

    _log("Reading Nexpay statement…")
    nx = pd.read_excel(nexpay_path, dtype=str)
    nx.columns = [c.strip() for c in nx.columns]

    se_nx = se.copy()
    _log(f"System rows total: {len(se_nx)}")

    # SE keys: Payment ID uppercased for case-insensitive match, External ID as integer string
    se_nx["_key_pid"] = se_nx["Payment ID"].str.strip().str.upper()

    def _ext_id_str(x) -> str:
        if pd.isna(x):
            return ""
        s = str(x).strip()
        if not s or s == "nan":
            return ""
        # Remove trailing .0 without float conversion (avoids precision loss on large ints)
        if '.' in s and 'e' not in s.lower() and 'E' not in s:
            s = s.split('.')[0]
        elif 'e' in s.lower():
            # Scientific notation — use Decimal to preserve precision
            try:
                from decimal import Decimal
                s = str(int(Decimal(s)))
            except Exception:
                return ""
        return s if s.lstrip('-').isdigit() else ""

    se_nx["_key_ext"] = se_nx["External ID"].apply(_ext_id_str)

    # Nexpay: split Details into numeric (→ External ID) and alphanumeric (→ Payment ID)
    nx["_det_token"] = nx["Details"].astype(str).str.split().str[0].str.strip()
    _is_numeric = nx["_det_token"].str.match(r"^\d+$")
    nx_alpha   = nx[~_is_numeric].copy()
    nx_numeric = nx[_is_numeric].copy()

    # Path A: alphanumeric → add "pay_" prefix → match SE Payment ID (uppercased)
    nx_alpha["_join_key"] = ("pay_" + nx_alpha["_det_token"]).str.upper()
    match_alpha = nx_alpha.merge(
        se_nx, left_on="_join_key", right_on="_key_pid", how="left", suffixes=("_nx", "_se")
    )
    match_alpha["_status"] = match_alpha["_key_pid"].notna().map(
        {True: STATUS_MATCHED, False: STATUS_NX_ONLY}
    )
    matched_pids = set(match_alpha.loc[match_alpha["_status"] == STATUS_MATCHED, "_key_pid"].dropna())

    # Path B: numeric → match SE External ID
    nx_numeric["_join_key"] = nx_numeric["_det_token"]
    match_numeric = nx_numeric.merge(
        se_nx, left_on="_join_key", right_on="_key_ext", how="left", suffixes=("_nx", "_se")
    )
    match_numeric["_status"] = match_numeric["_key_ext"].notna().map(
        {True: STATUS_MATCHED, False: STATUS_NX_ONLY}
    )

    # SE rows not matched by either path
    all_matched_se_pids = matched_pids | {
        r["_key_pid"] for _, r in match_numeric.iterrows()
        if r["_status"] == STATUS_MATCHED and pd.notna(r.get("_key_pid"))
    }
    se_only = se_nx[~se_nx["_key_pid"].isin(all_matched_se_pids)].copy()
    se_only["_status"] = STATUS_SE_ONLY

    # ── Combine all rows ─────────────────────────────────────────────────────
    _log("Merging…")
    merged = pd.concat([match_alpha, match_numeric, se_only], ignore_index=True, sort=False)

    # Amounts as float for totals / difference
    def _to_float(s) -> pd.Series:
        if not isinstance(s, pd.Series):
            return pd.Series(0.0, index=merged.index)
        return pd.to_numeric(s.astype(str).str.replace(",", ".", regex=False), errors="coerce").fillna(0.0)

    merged["_se_amt"] = _to_float(merged.get("Payment Amount"))
    merged["_se_fee"] = _to_float(merged.get("Fee"))
    merged["_nx_amt"] = _to_float(merged.get("Amount (EUR)"))
    merged["_diff"]   = merged["_se_amt"] - merged["_nx_amt"]

    se_status_col = merged.get("Payment status", pd.Series(dtype=str, index=merged.index))
    se_status_norm = se_status_col.astype(str).str.strip().str.lower()

    # ── Sheet 1: all matched, failed rows first ───────────────────────────────
    all_matched = merged[merged["_status"] == STATUS_MATCHED].copy()
    _ps_norm = all_matched.get("Payment status", pd.Series(dtype=str, index=all_matched.index))
    all_matched["_is_failed"] = _ps_norm.astype(str).str.strip().str.lower() == "failed"
    all_matched = all_matched.sort_values("_is_failed", ascending=False)

    # ── Sheet 2: SE processed rows with no NX match ───────────────────────────
    se_proc_nx_miss = merged[
        (merged["_status"] == STATUS_SE_ONLY) &
        (se_status_norm == "processed")
    ].copy()

    # ── Sheet 3: matched but SE status not processed ──────────────────────────
    matched_exc = merged[
        (merged["_status"] == STATUS_MATCHED) &
        (se_status_norm != "processed")
    ].copy()

    # ── Sheet 4: all unmatched (SE_ONLY + NX_ONLY) ───────────────────────────
    se_only_all = merged[merged["_status"] == STATUS_SE_ONLY].copy()
    nx_only_all = merged[merged["_status"] == STATUS_NX_ONLY].copy()

    # ── Stats ────────────────────────────────────────────────────────────────
    n_matched      = len(all_matched)
    n_matched_exc  = len(matched_exc)
    n_se_proc_miss = len(se_proc_nx_miss)
    n_se_only      = int((merged["_status"] == STATUS_SE_ONLY).sum())
    n_nx_only      = int((merged["_status"] == STATUS_NX_ONLY).sum())
    n_unmatched    = n_se_only + n_nx_only
    total_se       = merged.loc[merged["_status"] != STATUS_NX_ONLY, "_se_amt"].sum()
    total_nx       = merged.loc[merged["_status"] != STATUS_SE_ONLY, "_nx_amt"].sum()
    _log(
        f"Matched: {n_matched}  |  "
        f"SE processed / NX missing: {n_se_proc_miss}  |  "
        f"Exceptions: {n_matched_exc}  |  "
        f"Unmatched: {n_unmatched} (SE {n_se_only} + NX {n_nx_only})"
    )

    # ── Write Excel ──────────────────────────────────────────────────────────
    _log("Writing output…")
    wb = Workbook()
    hfont = header_font()
    dfont = data_font()

    def _v(row, col_name: str):
        val = row.get(col_name, "")
        return "" if pd.isna(val) else str(val)

    def _write_headers(ws, cols: list[tuple[str, int]]) -> None:
        for ci, (hdr, width) in enumerate(cols, start=1):
            cell = ws.cell(row=1, column=ci, value=hdr)
            cell.fill = DARK_BLUE
            cell.font = hfont
            cell.border = BORDER
            cell.alignment = Alignment(horizontal="center", vertical="center")
            ws.column_dimensions[cell.column_letter].width = width
        ws.row_dimensions[1].height = 20
        ws.freeze_panes = "A2"

    def _write_matched_rows(ws, df: pd.DataFrame, row_fill, exc_fill=None) -> None:
        for ri, (_, row) in enumerate(df.iterrows(), start=2):
            se_status = _v(row, "Payment status")
            alt = ALT if ri % 2 == 0 else WHITE

            values = [
                se_status,
                _v(row, "Payment ID"),
                _v(row, "Customer Name"),
                row["_se_amt"] if row["_se_amt"] != 0 else None,
                row["_se_fee"] if row["_se_fee"] != 0 else None,
                _v(row, "Date of Final Status Update"),
                _v(row, "Creation Date"),
                _v(row, "Processing Date"),
                _v(row, "Beneficiary / Sender"),
                _v(row, "Account number"),
                row["_nx_amt"] if row["_nx_amt"] != 0 else None,
                _v(row, "Payment number"),
                row["_diff"],
            ]

            status_fill = exc_fill if (exc_fill and se_status.lower().strip() != "processed") else row_fill

            for ci, (val, (col_name, _)) in enumerate(zip(values, MATCHED_COLS), start=1):
                cell = ws.cell(row=ri, column=ci, value=val)
                cell.font = dfont
                cell.border = BORDER
                cell.fill = status_fill if ci == 1 else alt
                if col_name in AMOUNT_COLS and val is not None:
                    cell.number_format = NUMBER_FMT
                    cell.alignment = Alignment(horizontal="right", vertical="center")
                else:
                    cell.alignment = Alignment(horizontal="left", vertical="center")

    def _write_se_proc_rows(ws, df: pd.DataFrame) -> None:
        for ri, (_, row) in enumerate(df.iterrows(), start=2):
            alt = ALT if ri % 2 == 0 else WHITE
            values = [
                _v(row, "Payment ID"),
                _v(row, "Customer Name"),
                _v(row, "Payment status"),
                row["_se_amt"] if row["_se_amt"] != 0 else None,
                row["_se_fee"] if row["_se_fee"] != 0 else None,
                _v(row, "Date of Final Status Update"),
            ]
            for ci, (val, (col_name, _)) in enumerate(zip(values, SE_PROC_COLS), start=1):
                cell = ws.cell(row=ri, column=ci, value=val)
                cell.font = dfont
                cell.border = BORDER
                cell.fill = alt
                if col_name in AMOUNT_COLS and val is not None:
                    cell.number_format = NUMBER_FMT
                    cell.alignment = Alignment(horizontal="right", vertical="center")
                else:
                    cell.alignment = Alignment(horizontal="left", vertical="center")

    def _write_nx_proc_rows(ws, df: pd.DataFrame) -> None:
        for ri, (_, row) in enumerate(df.iterrows(), start=2):
            alt = ALT if ri % 2 == 0 else WHITE
            values = [
                _v(row, "Creation Date"),
                _v(row, "Processing Date"),
                _v(row, "Beneficiary / Sender"),
                _v(row, "Account number"),
                row["_nx_amt"] if row["_nx_amt"] != 0 else None,
                _v(row, "Payment number"),
                _v(row, "Details"),
            ]
            for ci, (val, (col_name, _)) in enumerate(zip(values, NX_PROC_COLS), start=1):
                cell = ws.cell(row=ri, column=ci, value=val)
                cell.font = dfont
                cell.border = BORDER
                cell.fill = alt
                if col_name in AMOUNT_COLS and val is not None:
                    cell.number_format = NUMBER_FMT
                    cell.alignment = Alignment(horizontal="right", vertical="center")
                else:
                    cell.alignment = Alignment(horizontal="left", vertical="center")

    def _write_unmatched_rows(ws, se_df: pd.DataFrame, nx_df: pd.DataFrame) -> None:
        ri = 2
        for _, row in se_df.iterrows():
            alt = ALT if ri % 2 == 0 else WHITE
            values = [
                "SaltEdge",
                _v(row, "Payment status"),
                _v(row, "Payment ID"),
                _v(row, "Customer Name"),
                row["_se_amt"] if row["_se_amt"] != 0 else None,
                _v(row, "Date of Final Status Update"),
                "",
            ]
            for ci, (val, (col_name, _)) in enumerate(zip(values, UNMATCHED_COLS), start=1):
                cell = ws.cell(row=ri, column=ci, value=val)
                cell.font = dfont
                cell.border = BORDER
                cell.fill = alt
                if col_name in AMOUNT_COLS and val is not None:
                    cell.number_format = NUMBER_FMT
                    cell.alignment = Alignment(horizontal="right", vertical="center")
                else:
                    cell.alignment = Alignment(horizontal="left", vertical="center")
            ri += 1
        for _, row in nx_df.iterrows():
            alt = ALT if ri % 2 == 0 else WHITE
            values = [
                "Nexpay",
                "",
                _v(row, "Payment number"),
                _v(row, "Beneficiary / Sender"),
                row["_nx_amt"] if row["_nx_amt"] != 0 else None,
                _v(row, "Creation Date"),
                _v(row, "Details"),
            ]
            for ci, (val, (col_name, _)) in enumerate(zip(values, UNMATCHED_COLS), start=1):
                cell = ws.cell(row=ri, column=ci, value=val)
                cell.font = dfont
                cell.border = BORDER
                cell.fill = alt
                if col_name in AMOUNT_COLS and val is not None:
                    cell.number_format = NUMBER_FMT
                    cell.alignment = Alignment(horizontal="right", vertical="center")
                else:
                    cell.alignment = Alignment(horizontal="left", vertical="center")
            ri += 1

    # Sheet 1 — All matched (failed first)
    ws1 = wb.active
    ws1.title = "Matched"
    _write_headers(ws1, MATCHED_COLS)
    _write_matched_rows(ws1, all_matched, GREEN_FILL, exc_fill=ORANGE_FILL)

    # Sheet 2 — SE processed / NX missing
    ws2 = wb.create_sheet("SE Processed – NX Missing")
    _write_headers(ws2, SE_PROC_COLS)
    _write_se_proc_rows(ws2, se_proc_nx_miss)

    # Sheet 3 — NX processed / SE missing
    ws3 = wb.create_sheet("NX Processed – SE Missing")
    _write_headers(ws3, NX_PROC_COLS)
    _write_nx_proc_rows(ws3, nx_only_all)

    # Sheet 4 — Matched exceptions (non-processed)
    ws4 = wb.create_sheet("Exceptions")
    _write_headers(ws4, MATCHED_COLS)
    _write_matched_rows(ws4, matched_exc, ORANGE_FILL)

    # Sheet 5 — All unmatched (SE + NX)
    ws5 = wb.create_sheet("Unmatched")
    _write_headers(ws5, UNMATCHED_COLS)
    _write_unmatched_rows(ws5, se_only_all, nx_only_all)

    # ── Summary sheet ────────────────────────────────────────────────────────
    ws6 = wb.create_sheet("Summary")
    ws6.freeze_panes = "A2"
    for ci, (hdr, width) in enumerate([("Metric", 32), ("Value", 18)], start=1):
        cell = ws6.cell(row=1, column=ci, value=hdr)
        cell.fill = DARK_BLUE
        cell.font = hfont
        cell.border = BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws6.column_dimensions[cell.column_letter].width = width

    summary_rows = [
        ("Matched (all)",              n_matched),
        ("  — of which exceptions",    n_matched_exc),
        ("SE processed / NX missing",  n_se_proc_miss),
        ("NX processed / SE missing",  n_nx_only),
        ("Unmatched (SE + NX total)",  n_unmatched),
        ("  — SaltEdge only",          n_se_only),
        ("  — Nexpay only",            n_nx_only),
        ("Total SE Amount (EUR)",      round(total_se, 2)),
        ("Total NX Amount (EUR)",      round(total_nx, 2)),
        ("Difference",                 round(total_se - total_nx, 2)),
    ]
    for ri, (label, val) in enumerate(summary_rows, start=2):
        style_cell(ws6, ri, 1, label, fill=ALT if ri % 2 == 0 else WHITE, font=dfont)
        is_num = isinstance(val, float)
        style_cell(ws6, ri, 2, val,
                   fill=ALT if ri % 2 == 0 else WHITE,
                   font=dfont,
                   align="right",
                   number_format=NUMBER_FMT if is_num else None)

    # ── Save ─────────────────────────────────────────────────────────────────
    from pathlib import Path as _Path
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    _out_dir = _Path(out_dir)
    _out_dir.mkdir(parents=True, exist_ok=True)
    out_path = str(_out_dir.resolve() / f"nexpay_recon_{ts}.xlsx")
    wb.save(out_path)
    _log(f"Saved → {out_path}")

    return NexpayResult(
        matched=n_matched,
        se_processed_nx_missing=n_se_proc_miss,
        nx_processed_se_missing=n_nx_only,
        matched_exc=n_matched_exc,
        unmatched=n_unmatched,
        total_se_amount=round(total_se, 2),
        total_nx_amount=round(total_nx, 2),
        out_path=out_path,
    )
