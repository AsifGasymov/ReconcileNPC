"""Shared openpyxl style primitives for output xlsx files."""
from __future__ import annotations

from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


DARK_BLUE = PatternFill("solid", start_color="1F3864")
MID_BLUE = PatternFill("solid", start_color="2E75B6")
ALT = PatternFill("solid", start_color="EBF3FB")
WHITE = PatternFill("solid", start_color="FFFFFF")
GOLD = PatternFill("solid", start_color="FFF2CC")
GOLD_ALT = PatternFill("solid", start_color="FEF9E7")
GREEN_FILL = PatternFill("solid", start_color="E2EFDA")
ORANGE_FILL = PatternFill("solid", start_color="FCE4D6")
ORANGE_ALT = PatternFill("solid", start_color="F9D0BB")

THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def header_font(size: int = 9, color: str = "FFFFFF", bold: bool = True) -> Font:
    return Font(name="Arial", size=size, bold=bold, color=color)


def data_font(bold: bool = False) -> Font:
    return Font(name="Arial", size=9, bold=bold)


def style_cell(ws, row: int, col: int, value, *,
               fill: PatternFill | None = None,
               font: Font | None = None,
               align: str = "left",
               number_format: str | None = None):
    cell = ws.cell(row=row, column=col, value=value)
    if fill:
        cell.fill = fill
    if font:
        cell.font = font
    cell.border = BORDER
    cell.alignment = Alignment(horizontal=align, vertical="center")
    if number_format:
        cell.number_format = number_format
    return cell
