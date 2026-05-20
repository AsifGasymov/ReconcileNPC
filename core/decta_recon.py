"""DCT Stage 1 — Reconciliation: Main × Registry × Providers → New Transactions."""
from __future__ import annotations

from dataclasses import dataclass
from datetime import datetime, timedelta
from pathlib import Path
from typing import Callable, Optional

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


LogFn = Callable[[str], None]


@dataclass
class DectaReconResult:
    new_rows: int
    skipped: int
    dct_rows: int
    providers: int
    out_path: str


# ─── Styles ───────────────────────────────────────────────────────────────────
HDR_FILL  = PatternFill("solid", start_color="1E293B")
HDR_FONT  = Font(name="Calibri", size=9, bold=True, color="F8FAFC")
ROW_ODD   = PatternFill("solid", start_color="FFFFFF")
ROW_EVEN  = PatternFill("solid", start_color="F8FAFC")
ACC_ODD   = PatternFill("solid", start_color="F0FDF4")
ACC_EVEN  = PatternFill("solid", start_color="DCFCE7")
ACC_HDR   = PatternFill("solid", start_color="166534")
ACC_HFONT = Font(name="Calibri", size=9, bold=True, color="FFFFFF")

HAIR   = Side(style="hair", color="E2E8F0")
BORDER = Border(left=HAIR, right=HAIR, top=HAIR, bottom=HAIR)

DATA_FONT  = Font(name="Calibri", size=9, color="1E293B")
DATA_MUTED = Font(name="Calibri", size=9, color="64748B")
ACC_GREEN  = Font(name="Calibri", size=9, bold=True, color="166534")
ACC_RED    = Font(name="Calibri", size=9, bold=True, color="991B1B")

ADDED_COLS  = {"Transaction date +1", "Merchant Name", "Descriptor",
               "FTD/TD", "Provider currency", "Amount"}
RIGHT_COLS  = {"Processed Amount", "Processed Fee", "Amount"}
CENTER_COLS = {"Currency", "Provider currency", "FTD/TD"}
DROP_COLS   = {"CPI ID", "Status", "Issuer Country",
               "Merchant Account ID", "Merchant Account Name"}

DCT_COLS = [
    "ARN", "Payment ID", "Merchant Name", "Descriptor", "Descriptor city",
    "Terminal ID", "MCC", "Subtotal (Without VAT)", "Total", "Currency",
    "Type", "Status", "Creation Date", "Transaction Date", "Settlement Date",
    "Processing Date", "RRN", "Product name or description",
]

COL_WIDTHS = {
    "Transaction date +1": 18, "Payment ID": 30, "Merchant Name": 22,
    "Descriptor": 18, "FTD/TD": 8, "Card Mask": 18,
    "Date of Final Status": 16, "Commerce Account ID": 20,
    "Commerce Account Name": 22, "Provider": 12, "Method": 14,
    "Card Network": 12, "Currency": 9, "Provider currency": 14,
    "Processed Amount": 15, "Processed Fee": 13, "Amount": 13,
    "ARN": 26, "RRN": 18, "Product name or description": 38,
    "Subtotal (Without VAT)": 18, "Total": 12,
    "Creation Date": 14, "Transaction Date": 14,
    "Settlement Date": 14, "Processing Date": 14,
}


# ─── Readers ──────────────────────────────────────────────────────────────────

def _read_main(path: str) -> pd.DataFrame:
    df = pd.read_excel(path, dtype=str)
    df = df.dropna(how="all")
    df.columns = [str(c).strip() for c in df.columns]
    return df


def _read_registry(path: str) -> pd.DataFrame:
    p = path.lower()
    if p.endswith(".numbers"):
        try:
            from numbers_parser import Document
            doc = Document(path)
            table = doc.sheets[0].tables[0]
            rows = list(table.iter_rows())
            headers = [str(c.value).strip() for c in rows[0]]
            data = []
            for row in rows[1:]:
                vals = [str(c.value) if c.value is not None else None for c in row]
                if any(v for v in vals):
                    data.append(vals)
            df = pd.DataFrame(data, columns=headers)
        except ImportError:
            raise RuntimeError(
                ".numbers format requires 'numbers-parser'.\n"
                "Run: pip install numbers-parser"
            )
    elif p.endswith(".csv"):
        try:
            df = pd.read_csv(path, dtype=str, encoding="utf-8")
        except UnicodeDecodeError:
            df = pd.read_csv(path, dtype=str, encoding="latin-1")
    else:
        df = pd.read_excel(path, dtype=str)
    df = df.dropna(how="all")
    df.columns = [str(c).strip() for c in df.columns]
    return df


def _read_provider(path: str) -> pd.DataFrame:
    p = path.lower()
    if p.endswith(".csv"):
        try:
            df = pd.read_csv(path, dtype=str, encoding="utf-8")
        except UnicodeDecodeError:
            df = pd.read_csv(path, dtype=str, encoding="latin-1")
    elif p.endswith(".xls"):
        df = pd.read_excel(path, dtype=str, engine="xlrd")
    else:
        import openpyxl as _ox
        wb = _ox.load_workbook(path, read_only=True)
        sheet_name = "Orders" if "Orders" in wb.sheetnames else wb.sheetnames[0]
        wb.close()
        df = pd.read_excel(path, sheet_name=sheet_name, dtype=str)
    df = df.dropna(how="all")
    df.columns = [str(c).strip() for c in df.columns]
    if "Product name or description" in df.columns:
        df["Payment ID"] = df["Product name or description"].apply(
            lambda x: str(x)[:28].strip() if pd.notna(x) and x != "nan" else "")
    else:
        df["Payment ID"] = ""
    return df


# ─── Helpers ──────────────────────────────────────────────────────────────────

def _fmt_num(x) -> str:
    if not x or str(x) in ("", "nan", "None"):
        return ""
    try:
        return f"{float(str(x).replace(',', '.')):.2f}".replace(".", ",")
    except (ValueError, TypeError):
        return str(x).replace(".", ",")


def _sc(ws, row: int, col: int, val, *, fill=None, font=None,
        align: str = "left", fmt: str | None = None):
    cell = ws.cell(row=row, column=col, value=val)
    if fill:
        cell.fill = fill
    if font:
        cell.font = font
    cell.border = BORDER
    cell.alignment = Alignment(horizontal=align, vertical="center")
    if fmt:
        cell.number_format = fmt
    return cell


# ─── Excel writer ─────────────────────────────────────────────────────────────

def _write_xlsx(main_df: pd.DataFrame, dct_df: pd.DataFrame, out_path: str) -> None:
    wb = Workbook()

    # Sheet 1: New Transactions
    ws1 = wb.active
    ws1.title = "New Transactions"
    ws1.row_dimensions[1].height = 20
    cols = main_df.columns.tolist()

    for ci, col in enumerate(cols, 1):
        is_acc = col in ADDED_COLS
        al = "right" if col in RIGHT_COLS else ("center" if col in CENTER_COLS else "left")
        _sc(ws1, 1, ci, col,
            fill=ACC_HDR if is_acc else HDR_FILL,
            font=ACC_HFONT if is_acc else HDR_FONT,
            align=al)
        ws1.column_dimensions[get_column_letter(ci)].width = COL_WIDTHS.get(
            col, max(10, min(len(col) + 3, 26)))

    ws1.freeze_panes = "A2"

    for ri, (_, row) in enumerate(main_df.iterrows(), 2):
        even = ri % 2 == 0
        for ci, col in enumerate(cols, 1):
            val = row.get(col)
            if val == "nan" or val is None:
                val = ""
            is_acc = col in ADDED_COLS
            al = "right" if col in RIGHT_COLS else ("center" if col in CENTER_COLS else "left")
            fill = (ACC_EVEN if even else ACC_ODD) if is_acc else (ROW_EVEN if even else ROW_ODD)
            font = (ACC_GREEN if col == "Amount" else
                    ACC_RED if col == "Processed Fee" else
                    DATA_FONT) if is_acc else (
                DATA_MUTED if col == "Payment ID" else DATA_FONT)
            _sc(ws1, ri, ci, val, fill=fill, font=font, align=al)

    # Sheet 2: DCT
    ws2 = wb.create_sheet("DCT")
    ws2.row_dimensions[1].height = 20

    if not dct_df.empty:
        dct_cols = dct_df.columns.tolist()
        TEXT_COLS = {"ARN", "RRN", "Terminal ID", "MCC"}
        NUM_RIGHT = {"Subtotal (Without VAT)", "Total"}

        for ci, col in enumerate(dct_cols, 1):
            al = "right" if col in NUM_RIGHT else ("center" if col in CENTER_COLS else "left")
            _sc(ws2, 1, ci, col, fill=HDR_FILL, font=HDR_FONT, align=al)
            ws2.column_dimensions[get_column_letter(ci)].width = COL_WIDTHS.get(
                col, max(10, min(len(col) + 3, 26)))

        ws2.freeze_panes = "A2"

        for ri, (_, row) in enumerate(dct_df.iterrows(), 2):
            even = ri % 2 == 0
            fill = ROW_EVEN if even else ROW_ODD
            for ci, col in enumerate(dct_cols, 1):
                val = row.get(col)
                if val == "nan" or val is None:
                    val = ""
                al = "right" if col in NUM_RIGHT else ("center" if col in CENTER_COLS else "left")
                fnt = DATA_MUTED if col in ("ARN", "RRN") else DATA_FONT
                cell = _sc(ws2, ri, ci, val, fill=fill, font=fnt, align=al)
                if col in TEXT_COLS and val:
                    cell.number_format = "@"
    else:
        ws2["A1"] = "No provider data loaded"

    wb.save(out_path)


# ─── Main pipeline ────────────────────────────────────────────────────────────

def run_decta_recon(*, main_path: str, registry_path: str,
                    provider_paths: list[str], out_dir: str,
                    log: Optional[LogFn] = None) -> DectaReconResult:
    def lg(msg: str) -> None:
        if log:
            log(msg)

    lg("Reading Main file…")
    main_df = _read_main(main_path)
    main_pids = set(main_df["Payment ID"].dropna().str.strip().tolist())
    lg(f"  Main: {len(main_df)} rows, {len(main_pids)} Payment IDs")

    lg("Reading Payment Registry…")
    reg_df = _read_registry(registry_path)
    lg(f"  Registry: {len(reg_df)} rows")

    lg("Filtering — removing IDs already in Main…")
    reg_df["Payment ID"] = reg_df["Payment ID"].str.strip()
    new_df = reg_df[~reg_df["Payment ID"].isin(main_pids)].copy()
    skipped = len(reg_df) - len(new_df)
    lg(f"  New rows: {len(new_df)} (skipped {skipped})")

    lg("Reading provider files…")
    provider_dfs = []
    for pp in provider_paths:
        try:
            pf = _read_provider(pp)
            provider_dfs.append(pf)
            lg(f"  {Path(pp).name}: {len(pf)} rows")
        except Exception as e:
            lg(f"  SKIP {Path(pp).name}: {e}")

    combined_provider = (pd.concat(provider_dfs, ignore_index=True)
                         if provider_dfs else pd.DataFrame())
    lg(f"  Total provider rows: {len(combined_provider)}")

    # Build lookup: Payment ID → enrichment fields
    prov_lookup: dict[str, dict] = {}
    if not combined_provider.empty:
        for _, row in combined_provider.iterrows():
            pid = str(row.get("Payment ID", "")).strip()
            if not pid:
                continue
            prov_lookup[pid] = {
                "Merchant Name":     str(row.get("Merchant Name", "") or ""),
                "Descriptor":        str(row.get("Descriptor", "") or ""),
                "Provider currency": str(row.get("Currency", "") or ""),
                "Amount":            _fmt_num(str(row.get("Total", "") or "")),
                "Transaction Date":  str(row.get("Transaction Date", "") or ""),
            }

    lg("Transforming columns…")
    for col in list(DROP_COLS):
        if col in new_df.columns:
            new_df.drop(columns=[col], inplace=True)

    new_df.insert(0, "Transaction date +1",
                  new_df["Payment ID"].map(
                      lambda x: prov_lookup.get(x, {}).get("Transaction Date", ""))
                  if prov_lookup else "")

    pid_idx = new_df.columns.tolist().index("Payment ID") + 1
    new_df.insert(pid_idx,     "Merchant Name",
                  new_df["Payment ID"].map(lambda x: prov_lookup.get(x, {}).get("Merchant Name", "")))
    new_df.insert(pid_idx + 1, "Descriptor",
                  new_df["Payment ID"].map(lambda x: prov_lookup.get(x, {}).get("Descriptor", "")))
    new_df.insert(pid_idx + 2, "FTD/TD", "")

    if "Currency" in new_df.columns:
        cur_idx = new_df.columns.tolist().index("Currency") + 1
        new_df.insert(cur_idx, "Provider currency",
                      new_df["Payment ID"].map(
                          lambda x: prov_lookup.get(x, {}).get("Provider currency", "")))
    else:
        new_df["Provider currency"] = ""

    for col in ("Processed Amount", "Processed Fee"):
        if col in new_df.columns:
            new_df[col] = new_df[col].apply(_fmt_num)

    new_df["Amount"] = new_df["Payment ID"].map(
        lambda x: prov_lookup.get(x, {}).get("Amount", ""))

    # Build DCT sheet
    dct_df = pd.DataFrame()
    if not combined_provider.empty:
        for col in DCT_COLS:
            if col not in combined_provider.columns:
                combined_provider[col] = ""
        dct_df = combined_provider[DCT_COLS].copy()
        for nc in ("Subtotal (Without VAT)", "Total"):
            if nc in dct_df.columns:
                dct_df[nc] = dct_df[nc].apply(_fmt_num)

    lg("Writing workbook…")
    out_name = f"DCT_Reconciliation_{datetime.now().strftime('%d-%m-%Y')}.xlsx"
    out_path = str(Path(out_dir) / out_name)
    _write_xlsx(new_df, dct_df, out_path)

    return DectaReconResult(
        new_rows=len(new_df),
        skipped=skipped,
        dct_rows=len(dct_df),
        providers=len(provider_paths),
        out_path=out_path,
    )
