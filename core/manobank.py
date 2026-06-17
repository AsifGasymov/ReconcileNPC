"""ManoBank × SaltEdge reconciliation report."""
from __future__ import annotations

import re
from dataclasses import dataclass
from datetime import datetime
from typing import Callable, Optional

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment

from .excel_styles import (
    ALT, BORDER, DARK_BLUE, GREEN_FILL, ORANGE_FILL, WHITE,
    data_font, header_font, style_cell,
)

LogFn = Callable[[str], None]

STATUS_MATCHED = "Matched"
STATUS_SE_ONLY = "SaltEdge only"
STATUS_MB_ONLY = "ManoBank only"

NUMBER_FMT = "#,##0.00"

# Sheets 1 & 4 — matched rows (both sides)
MB_MATCHED_COLS: list[tuple[str, int]] = [
    ("SE Status",          16),
    ("Payment ID",         32),
    ("Customer Name",      24),
    ("SE Amount (EUR)",    16),
    ("SE Fee",             12),
    ("SE Date",            14),
    ("MB Date",            14),
    ("MB Reference",       16),
    ("MB Payer",           28),
    ("MB Payer Account",   26),
    ("MB Amount (EUR)",    16),
    ("Difference",         14),
    ("Cost (1.7%)",        14),
    ("Fixed Fee",          12),
]

COST_RATE      = 0.017
MB_FIXED_FEE   = 0.25

# Sheet 2 — SE processed, no MB match
SE_PROC_COLS: list[tuple[str, int]] = [
    ("Payment ID",         32),
    ("Customer Name",      24),
    ("Payment status",     16),
    ("SE Amount (EUR)",    16),
    ("SE Fee",             12),
    ("SE Date",            16),
]

# Sheet 3 — MB rows, no SE match
MB_ONLY_COLS: list[tuple[str, int]] = [
    ("MB Date",            14),
    ("MB Reference",       16),
    ("MB Payer",           28),
    ("MB Payer Account",   26),
    ("MB Amount (EUR)",    16),
    ("Payment details",    50),
]

# Sheet 5 — all unmatched combined
MB_UNMATCHED_COLS: list[tuple[str, int]] = [
    ("Source",                  14),
    ("Status",                  16),
    ("Payment ID / MB Ref",     32),
    ("Customer / Payer",        28),
    ("Amount (EUR)",            16),
    ("Date",                    16),
    ("Details / Account",       40),
]

AMOUNT_COLS = {
    "SE Amount (EUR)", "MB Amount (EUR)", "SE Fee", "Difference",
    "Amount (EUR)", "Cost (1.7%)", "Fixed Fee",
}


@dataclass
class ManoBankResult:
    matched: int
    se_processed_mb_missing: int
    mb_processed_se_missing: int
    matched_exc: int
    unmatched: int
    total_se_amount: float
    total_mb_amount: float
    out_path: str


def _extract_token(details: str) -> str:
    """Extract first token from 'Payment details: <token>' in ManoBank details field."""
    m = re.search(r'Payment details:\s*(\S+)', str(details))
    return m.group(1).strip() if m else ""


def _ext_id_str(x) -> str:
    """Convert External ID to plain integer string without float precision loss."""
    if pd.isna(x):
        return ""
    s = str(x).strip()
    if not s or s == "nan":
        return ""
    if '.' in s and 'e' not in s.lower() and 'E' not in s:
        s = s.split('.')[0]
    elif 'e' in s.lower():
        try:
            from decimal import Decimal
            s = str(int(Decimal(s)))
        except Exception:
            return ""
    return s if s.lstrip('-').isdigit() else ""


def run_manobank(
    *,
    system_paths: list[str],
    manobank_path: str,
    out_dir: str,
    log: Optional[LogFn] = None,
) -> ManoBankResult:
    def _log(msg: str) -> None:
        if log:
            log(msg)

    _log(f"Reading {len(system_paths)} system export(s)…")
    parts = []
    for p in system_paths:
        df = pd.read_csv(p, dtype=str)
        df.columns = [c.strip() for c in df.columns]
        parts.append(df)
    se = pd.concat(parts, ignore_index=True) if len(parts) > 1 else parts[0]

    _log("Reading ManoBank statement…")
    mb = pd.read_excel(manobank_path, dtype=str, header=13)
    mb.columns = [c.strip() for c in mb.columns]
    mb = mb.dropna(subset=["Date"]).copy()
    _log(f"ManoBank rows: {len(mb)}  |  System rows: {len(se)}")

    # SE keys
    se_nx = se.copy()
    se_nx["_key_pid"] = se_nx["Payment ID"].str.strip().str.upper()
    se_nx["_key_ext"] = se_nx["External ID"].apply(_ext_id_str)

    # ManoBank: extract token from Payment details field
    mb["_det_token"] = mb["Payment details"].apply(_extract_token)
    _is_numeric = mb["_det_token"].str.match(r"^\d+$")
    mb_alpha   = mb[~_is_numeric].copy()
    mb_numeric = mb[_is_numeric].copy()

    # Path A: alphanumeric → pay_ prefix → SE Payment ID
    mb_alpha["_join_key"] = ("pay_" + mb_alpha["_det_token"]).str.upper()
    match_alpha = mb_alpha.merge(
        se_nx, left_on="_join_key", right_on="_key_pid", how="left", suffixes=("_mb", "_se")
    )
    match_alpha["_status"] = match_alpha["_key_pid"].notna().map(
        {True: STATUS_MATCHED, False: STATUS_MB_ONLY}
    )
    matched_pids = set(
        match_alpha.loc[match_alpha["_status"] == STATUS_MATCHED, "_key_pid"].dropna()
    )

    # Path B: numeric → SE External ID
    mb_numeric["_join_key"] = mb_numeric["_det_token"]
    match_numeric = mb_numeric.merge(
        se_nx, left_on="_join_key", right_on="_key_ext", how="left", suffixes=("_mb", "_se")
    )
    match_numeric["_status"] = match_numeric["_key_ext"].notna().map(
        {True: STATUS_MATCHED, False: STATUS_MB_ONLY}
    )

    # SE rows not matched by either path
    all_matched_se_pids = matched_pids | {
        r["_key_pid"] for _, r in match_numeric.iterrows()
        if r["_status"] == STATUS_MATCHED and pd.notna(r.get("_key_pid"))
    }
    se_only = se_nx[~se_nx["_key_pid"].isin(all_matched_se_pids)].copy()
    se_only["_status"] = STATUS_SE_ONLY

    # ── Combine ──────────────────────────────────────────────────────────────
    _log("Merging…")
    merged = pd.concat([match_alpha, match_numeric, se_only], ignore_index=True, sort=False)

    def _to_float(s) -> pd.Series:
        if not isinstance(s, pd.Series):
            return pd.Series(0.0, index=merged.index)
        return pd.to_numeric(
            s.astype(str).str.replace(",", ".", regex=False), errors="coerce"
        ).fillna(0.0)

    merged["_se_amt"] = _to_float(merged.get("Payment Amount"))
    merged["_se_fee"] = _to_float(merged.get("Fee"))
    merged["_mb_amt"] = _to_float(merged.get("Amount"))
    merged["_diff"]   = merged["_se_amt"] - merged["_mb_amt"]

    se_status_norm = merged.get(
        "Payment status", pd.Series(dtype=str, index=merged.index)
    ).astype(str).str.strip().str.lower()

    # Groups
    all_matched = merged[
        (merged["_status"] == STATUS_MATCHED) &
        (se_status_norm == "processed")
    ].copy()

    se_proc_mb_miss = merged[
        (merged["_status"] == STATUS_SE_ONLY) & (se_status_norm == "processed")
    ].copy()

    mb_only_all = merged[merged["_status"] == STATUS_MB_ONLY].copy()

    matched_exc = merged[
        (merged["_status"] == STATUS_MATCHED) & (se_status_norm != "processed")
    ].copy()

    se_only_all = merged[merged["_status"] == STATUS_SE_ONLY].copy()

    # Stats
    n_matched      = len(all_matched)
    n_matched_exc  = len(matched_exc)
    n_se_proc_miss = len(se_proc_mb_miss)
    n_mb_only      = int((merged["_status"] == STATUS_MB_ONLY).sum())
    n_se_only_cnt  = int((merged["_status"] == STATUS_SE_ONLY).sum())
    n_unmatched    = n_se_only_cnt + n_mb_only
    total_se       = merged.loc[merged["_status"] != STATUS_MB_ONLY, "_se_amt"].sum()
    total_mb       = merged.loc[merged["_status"] != STATUS_SE_ONLY, "_mb_amt"].sum()
    _log(
        f"Matched: {n_matched}  |  SE proc / MB missing: {n_se_proc_miss}  |  "
        f"MB / SE missing: {n_mb_only}  |  Exceptions: {n_matched_exc}  |  "
        f"Unmatched: {n_unmatched}"
    )

    # ── Write Excel ──────────────────────────────────────────────────────────
    _log("Writing output…")
    wb = Workbook()
    hfont = header_font()
    dfont = data_font()

    def _v(row, col_name: str) -> str:
        val = row.get(col_name, "")
        return "" if pd.isna(val) else str(val)

    def _write_headers(ws, cols: list[tuple[str, int]]) -> None:
        for ci, (hdr, width) in enumerate(cols, start=1):
            cell = ws.cell(row=1, column=ci, value=hdr)
            cell.fill = DARK_BLUE
            cell.font = hfont
            cell.border = BORDER
            cell.alignment = Alignment(horizontal="center", vertical="center")
            ws.column_dimensions[cell.column_letter].width = width
        ws.row_dimensions[1].height = 20
        ws.freeze_panes = "A2"

    def _write_matched_rows(ws, df: pd.DataFrame, row_fill, exc_fill=None) -> None:
        for ri, (_, row) in enumerate(df.iterrows(), start=2):
            se_status = _v(row, "Payment status")
            alt = ALT if ri % 2 == 0 else WHITE
            se_amt = row["_se_amt"]
            values = [
                se_status,
                _v(row, "Payment ID"),
                _v(row, "Customer Name"),
                se_amt if se_amt != 0 else None,
                row["_se_fee"] if row["_se_fee"] != 0 else None,
                _v(row, "Date of Final Status Update"),
                _v(row, "Date"),
                _v(row, "Reference no"),
                _v(row, "Payer"),
                _v(row, "Payer account"),
                row["_mb_amt"] if row["_mb_amt"] != 0 else None,
                row["_diff"],
                round(se_amt * COST_RATE, 2) if se_amt != 0 else None,
                MB_FIXED_FEE,
            ]
            status_fill = (
                exc_fill if (exc_fill and se_status.lower().strip() != "processed")
                else row_fill
            )
            for ci, (val, (col_name, _)) in enumerate(zip(values, MB_MATCHED_COLS), start=1):
                cell = ws.cell(row=ri, column=ci, value=val)
                cell.font = dfont
                cell.border = BORDER
                cell.fill = status_fill if ci == 1 else alt
                if col_name in AMOUNT_COLS and val is not None:
                    cell.number_format = NUMBER_FMT
                    cell.alignment = Alignment(horizontal="right", vertical="center")
                else:
                    cell.alignment = Alignment(horizontal="left", vertical="center")

    def _write_se_proc_rows(ws, df: pd.DataFrame) -> None:
        for ri, (_, row) in enumerate(df.iterrows(), start=2):
            alt = ALT if ri % 2 == 0 else WHITE
            values = [
                _v(row, "Payment ID"),
                _v(row, "Customer Name"),
                _v(row, "Payment status"),
                row["_se_amt"] if row["_se_amt"] != 0 else None,
                row["_se_fee"] if row["_se_fee"] != 0 else None,
                _v(row, "Date of Final Status Update"),
            ]
            for ci, (val, (col_name, _)) in enumerate(zip(values, SE_PROC_COLS), start=1):
                cell = ws.cell(row=ri, column=ci, value=val)
                cell.font = dfont
                cell.border = BORDER
                cell.fill = alt
                if col_name in AMOUNT_COLS and val is not None:
                    cell.number_format = NUMBER_FMT
                    cell.alignment = Alignment(horizontal="right", vertical="center")
                else:
                    cell.alignment = Alignment(horizontal="left", vertical="center")

    def _write_mb_only_rows(ws, df: pd.DataFrame) -> None:
        for ri, (_, row) in enumerate(df.iterrows(), start=2):
            alt = ALT if ri % 2 == 0 else WHITE
            values = [
                _v(row, "Date"),
                _v(row, "Reference no"),
                _v(row, "Payer"),
                _v(row, "Payer account"),
                row["_mb_amt"] if row["_mb_amt"] != 0 else None,
                _v(row, "Payment details"),
            ]
            for ci, (val, (col_name, _)) in enumerate(zip(values, MB_ONLY_COLS), start=1):
                cell = ws.cell(row=ri, column=ci, value=val)
                cell.font = dfont
                cell.border = BORDER
                cell.fill = alt
                if col_name in AMOUNT_COLS and val is not None:
                    cell.number_format = NUMBER_FMT
                    cell.alignment = Alignment(horizontal="right", vertical="center")
                else:
                    cell.alignment = Alignment(horizontal="left", vertical="center")

    def _write_unmatched_rows(ws, se_df: pd.DataFrame, mb_df: pd.DataFrame) -> None:
        ri = 2
        for _, row in se_df.iterrows():
            alt = ALT if ri % 2 == 0 else WHITE
            values = [
                "SaltEdge",
                _v(row, "Payment status"),
                _v(row, "Payment ID"),
                _v(row, "Customer Name"),
                row["_se_amt"] if row["_se_amt"] != 0 else None,
                _v(row, "Date of Final Status Update"),
                "",
            ]
            for ci, (val, (col_name, _)) in enumerate(zip(values, MB_UNMATCHED_COLS), start=1):
                cell = ws.cell(row=ri, column=ci, value=val)
                cell.font = dfont
                cell.border = BORDER
                cell.fill = alt
                if col_name in AMOUNT_COLS and val is not None:
                    cell.number_format = NUMBER_FMT
                    cell.alignment = Alignment(horizontal="right", vertical="center")
                else:
                    cell.alignment = Alignment(horizontal="left", vertical="center")
            ri += 1
        for _, row in mb_df.iterrows():
            alt = ALT if ri % 2 == 0 else WHITE
            values = [
                "ManoBank",
                "",
                _v(row, "Reference no"),
                _v(row, "Payer"),
                row["_mb_amt"] if row["_mb_amt"] != 0 else None,
                _v(row, "Date"),
                _v(row, "Payer account"),
            ]
            for ci, (val, (col_name, _)) in enumerate(zip(values, MB_UNMATCHED_COLS), start=1):
                cell = ws.cell(row=ri, column=ci, value=val)
                cell.font = dfont
                cell.border = BORDER
                cell.fill = alt
                if col_name in AMOUNT_COLS and val is not None:
                    cell.number_format = NUMBER_FMT
                    cell.alignment = Alignment(horizontal="right", vertical="center")
                else:
                    cell.alignment = Alignment(horizontal="left", vertical="center")
            ri += 1

    # Sheet 1 — All matched (failed first)
    ws1 = wb.active
    ws1.title = "Matched"
    _write_headers(ws1, MB_MATCHED_COLS)
    _write_matched_rows(ws1, all_matched, GREEN_FILL, exc_fill=ORANGE_FILL)

    # Sheet 2 — SE processed / MB missing
    ws2 = wb.create_sheet("SE Processed – MB Missing")
    _write_headers(ws2, SE_PROC_COLS)
    _write_se_proc_rows(ws2, se_proc_mb_miss)

    # Sheet 3 — MB processed / SE missing
    ws3 = wb.create_sheet("MB Processed – SE Missing")
    _write_headers(ws3, MB_ONLY_COLS)
    _write_mb_only_rows(ws3, mb_only_all)

    # Sheet 4 — Matched exceptions
    ws4 = wb.create_sheet("Exceptions")
    _write_headers(ws4, MB_MATCHED_COLS)
    _write_matched_rows(ws4, matched_exc, ORANGE_FILL)

    # Sheet 5 — All unmatched
    ws5 = wb.create_sheet("Unmatched")
    _write_headers(ws5, MB_UNMATCHED_COLS)
    _write_unmatched_rows(ws5, se_only_all, mb_only_all)

    # Sheet 6 — Summary
    ws6 = wb.create_sheet("Summary")
    ws6.freeze_panes = "A2"
    for ci, (hdr, width) in enumerate([("Metric", 32), ("Value", 18)], start=1):
        cell = ws6.cell(row=1, column=ci, value=hdr)
        cell.fill = DARK_BLUE
        cell.font = hfont
        cell.border = BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws6.column_dimensions[cell.column_letter].width = width

    summary_rows = [
        ("Matched (all)",              n_matched),
        ("  — of which exceptions",    n_matched_exc),
        ("SE processed / MB missing",  n_se_proc_miss),
        ("MB processed / SE missing",  n_mb_only),
        ("Unmatched (SE + MB total)",  n_unmatched),
        ("  — SaltEdge only",          n_se_only_cnt),
        ("  — ManoBank only",          n_mb_only),
        ("Total SE Amount (EUR)",      round(total_se, 2)),
        ("Total MB Amount (EUR)",      round(total_mb, 2)),
        ("Difference",                 round(total_se - total_mb, 2)),
    ]
    for ri, (label, val) in enumerate(summary_rows, start=2):
        style_cell(ws6, ri, 1, label, fill=ALT if ri % 2 == 0 else WHITE, font=dfont)
        is_num = isinstance(val, float)
        style_cell(ws6, ri, 2, val,
                   fill=ALT if ri % 2 == 0 else WHITE,
                   font=dfont,
                   align="right",
                   number_format=NUMBER_FMT if is_num else None)

    # Save
    from pathlib import Path as _Path
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    _out_dir = _Path(out_dir)
    _out_dir.mkdir(parents=True, exist_ok=True)
    out_path = str(_out_dir.resolve() / f"manobank_recon_{ts}.xlsx")
    wb.save(out_path)
    _log(f"Saved → {out_path}")

    return ManoBankResult(
        matched=n_matched,
        se_processed_mb_missing=n_se_proc_miss,
        mb_processed_se_missing=n_mb_only,
        matched_exc=n_matched_exc,
        unmatched=n_unmatched,
        total_se_amount=round(total_se, 2),
        total_mb_amount=round(total_mb, 2),
        out_path=out_path,
    )
