"""Stage 1 — Reconciliation: Grafana × CDQ × General → SPNT/CDQ workbook."""
from __future__ import annotations

import re
import tempfile
from dataclasses import dataclass
from datetime import datetime
from typing import Callable, Iterable, Optional

import pandas as pd
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from .excel_styles import (
    ALT, GOLD, GOLD_ALT, MID_BLUE, WHITE,
    data_font, header_font, style_cell,
)
from .io import read_table


LogFn = Callable[[str], None]


@dataclass
class ReconResult:
    sheet1_rows: int
    sheet2_rows: int
    excluded: int
    out_path: str


PROV_COLS_ORDER = [
    ("Payment ID", "Payment ID"),
    ("Merchant Name", "Merchant Name"),
    ("Subtotal (Without VAT)", "Subtotal (Without VAT)"),
    ("Total", "Total"),
    ("Currency", "Currency"),
    ("Type", "Type"),
    ("Status", "Status"),
    ("Creation Date", "Creation Date"),
    ("Creation Time", "Creation Time"),
    ("Transaction Date", "Transaction Date"),
    ("Transaction Time", "Transaction Time"),
    ("Settlement Date", "Settlement Date"),
    ("Settlement Time", "Settlement Time"),
    ("Processing Date", "Processing Date"),
    ("Processing Time", "Processing Time"),
    ("ARN", "ARN"),
    ("RRN", "RRN"),
    ("Product name or description", "Product name or description"),
]

SHEET1_COLS = [
    ("Merchant Account ID", "Merchant Account ID"),
    ("Payment ID", "Payment ID"),
    ("Status", "Status"),
    ("Card Mask", "Card Mask"),
    ("Date of Final Status", "Date of Final Status"),
    ("Commerce Account ID", "Commerce Account ID"),
    ("Commerce Account Name", "Commerce Account Name"),
    ("Provider", "Provider"),
    ("Method", "Method"),
    ("Card Network", "Card Network"),
    ("Currency", "Currency"),
    ("Type", "Type"),
    ("Processed Amount", "Processed Amount"),
    ("Processed Fee", "Processed Fee"),
    ("Amount", "Amount"),
    ("Company", "Company"),
]

ADDED_COLS = {"Type", "Amount", "Company"}

NUMERIC_COLS = {
    "Subtotal (Without VAT)", "Total",
    "Processed Amount", "Processed Fee", "Amount",
}

DATE_COLS = {
    "Date of Final Status", "Creation Date", "Transaction Date",
    "Settlement Date", "Processing Date",
}

COL_WIDTHS = {
    "Merchant Account ID": 20, "Payment ID": 30, "Status": 12,
    "Card Mask": 18, "Date of Final Status": 16,
    "Commerce Account ID": 20, "Commerce Account Name": 22,
    "Provider": 12, "Method": 14, "Card Network": 14,
    "Currency": 10, "Type": 12, "Processed Amount": 16,
    "Processed Fee": 14, "Amount": 14, "Company": 20,
    "Merchant Name": 20, "Subtotal (Without VAT)": 18,
    "Total": 12, "ARN": 26, "RRN": 18,
    "Creation Date": 14, "Transaction Date": 14,
    "Settlement Date": 14, "Processing Date": 14,
    "Product name or description": 40,
}


def _read_grafana(path: str) -> pd.DataFrame:
    return read_table(path)


def _read_provsystems(path: str) -> tuple[pd.DataFrame, pd.DataFrame]:
    p = path.lower()
    if p.endswith(".csv") or p.endswith(".xls"):
        df = read_table(path)
    else:
        df = read_table(path, sheet_name="Orders")

    prod_col = "Product name or description"
    df["Payment ID"] = df[prod_col].apply(
        lambda x: str(x)[:28].strip() if pd.notna(x) and x != "nan" else ""
    )

    lookup = df[["Payment ID", "Currency", "Total", "Merchant Name"]].copy()
    lookup = lookup.rename(columns={
        "Currency": "_prov_currency",
        "Total": "_prov_total",
        "Merchant Name": "_prov_merchant",
    })

    sheet2_cols, sheet2_names = [], []
    for src, out in PROV_COLS_ORDER:
        if src in df.columns:
            sheet2_cols.append(src)
            sheet2_names.append(out)
        elif src == "Payment ID":
            sheet2_cols.append("Payment ID")
            sheet2_names.append(out)

    sheet2 = df[sheet2_cols].copy()
    sheet2.columns = sheet2_names
    return sheet2, lookup


def _read_general(path: str) -> set[str]:
    df = read_table(path)
    col = df.columns[0]
    return set(df[col].dropna().str.strip().tolist())


def _merge_provsystems(paths: Iterable[str]) -> str:
    """Concatenate multiple CDQ files into a single temp xlsx."""
    paths = list(paths)
    if len(paths) == 1:
        return paths[0]

    dfs = []
    for pp in paths:
        p = pp.lower()
        if p.endswith(".csv") or p.endswith(".xls"):
            df = read_table(pp)
        else:
            df = read_table(pp, sheet_name="Orders")
        dfs.append(df)

    combined = pd.concat(dfs, ignore_index=True)
    tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
    combined.to_excel(tmp.name, sheet_name="Orders", index=False)
    return tmp.name


def _fmt_num(val) -> str:
    if val is None or str(val) in ("", "nan", "None"):
        return ""
    try:
        return f"{float(str(val).replace(',', '.')):.2f}".replace(".", ",")
    except (ValueError, TypeError):
        return str(val)


def _fmt_date(val) -> str:
    if not val or str(val) in ("nan", "None", ""):
        return ""
    s = str(val)
    if re.match(r"\d{2}\.\d{2}\.\d{4}$", s):
        return s
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%d.%m.%Y %H:%M:%S"):
        try:
            return datetime.strptime(s[:19], fmt).strftime("%d.%m.%Y")
        except ValueError:
            pass
    return s[:10]


def _write_sheet(wb: Workbook, df: pd.DataFrame, sheet_name: str,
                 added_cols: set[str] | None = None) -> None:
    added_cols = added_cols or set()
    ws = wb.create_sheet(sheet_name)
    ws.row_dimensions[1].height = 18

    cols = df.columns.tolist()
    for ci, col in enumerate(cols, 1):
        is_add = col in added_cols
        fill = GOLD if is_add else MID_BLUE
        font_color = "000000" if is_add else "FFFFFF"
        style_cell(ws, 1, ci, col, fill=fill, font=header_font(color=font_color), align="center")
        ws.column_dimensions[get_column_letter(ci)].width = COL_WIDTHS.get(
            col, max(10, min(len(col) + 3, 25))
        )

    ws.freeze_panes = "A2"

    for ri, (_, row) in enumerate(df.iterrows(), 2):
        even = ri % 2 == 0
        for ci, col in enumerate(cols, 1):
            val = row.get(col)
            if val == "nan" or val is None:
                val = ""
            if col in DATE_COLS:
                val = _fmt_date(val)
            is_add = col in added_cols
            if is_add:
                fill = GOLD if even else GOLD_ALT
            else:
                fill = ALT if even else WHITE
            style_cell(ws, ri, ci, val, fill=fill, font=data_font())


def _write_workbook(sheet1: pd.DataFrame, sheet2: pd.DataFrame, out_path: str) -> None:
    for col in NUMERIC_COLS:
        if col in sheet1.columns:
            sheet1[col] = sheet1[col].apply(_fmt_num)
        if col in sheet2.columns:
            sheet2[col] = sheet2[col].apply(_fmt_num)

    wb = Workbook()
    wb.remove(wb.active)
    _write_sheet(wb, sheet1, "SPNT", added_cols=ADDED_COLS)
    _write_sheet(wb, sheet2, "CDQ")
    wb.save(out_path)


def run_reconciliation(*, grafana_path: str, cdq_paths: list[str],
                       general_path: str, out_dir: str,
                       log: Optional[LogFn] = None) -> ReconResult:
    """Execute the reconciliation pipeline and write the workbook."""
    def lg(msg: str) -> None:
        if log:
            log(msg)

    lg("Reading Grafana file…")
    test_df = _read_grafana(grafana_path)
    lg(f"  rows: {len(test_df)}")

    lg("Merging CDQ files…")
    prov_path = _merge_provsystems(cdq_paths)
    sheet2_df, prov_lookup = _read_provsystems(prov_path)
    lg(f"  CDQ rows: {len(sheet2_df)}")

    lg("Reading General (exclusions)…")
    exclude_ids = _read_general(general_path)
    lg(f"  exclusion IDs: {len(exclude_ids)}")

    lg("Filtering by General…")
    test_df["Payment ID"] = test_df["Payment ID"].str.strip()
    before = len(test_df)
    test_df = test_df[~test_df["Payment ID"].isin(exclude_ids)].copy()
    lg(f"  removed: {before - len(test_df)}, kept: {len(test_df)}")

    lg("Joining CDQ data…")
    prov_lookup = prov_lookup.drop_duplicates(subset="Payment ID")
    test_df = test_df.merge(prov_lookup, on="Payment ID", how="left")

    lg("Building Sheet 1…")
    test_df["Type"] = test_df["_prov_currency"].fillna("")
    test_df["Amount"] = test_df["_prov_total"].fillna("")
    test_df["Company"] = test_df["_prov_merchant"].fillna("")

    sheet1_cols, sheet1_names = [], []
    for src, out in SHEET1_COLS:
        if src in test_df.columns:
            sheet1_cols.append(src)
            sheet1_names.append(out)
    sheet1 = test_df[sheet1_cols].copy()
    sheet1.columns = sheet1_names

    out_name = f"Reconciliation_{datetime.now().strftime('%d-%m-%Y')}.xlsx"
    from pathlib import Path
    out_path = str(Path(out_dir) / out_name)

    lg("Writing workbook…")
    _write_workbook(sheet1, sheet2_df, out_path)

    return ReconResult(
        sheet1_rows=len(sheet1),
        sheet2_rows=len(sheet2_df),
        excluded=before - len(test_df),
        out_path=out_path,
    )
